from __future__ import annotations

import io
import multiprocessing
import os
import queue
import re
import shutil
import ssl
import stat
import subprocess
import sys
import tempfile
import threading
import time
import urllib.error
import urllib.parse
import urllib.request
import webbrowser
from dataclasses import dataclass
from pathlib import Path
from typing import Callable, Optional
from zipfile import BadZipFile
import json
import xml.etree.ElementTree as ET

try:
    import certifi
    import cv2
    import paramiko
    from PIL import Image, ImageDraw, ImageFont, ImageTk
    from openpyxl import Workbook, load_workbook
    from openpyxl.drawing.image import Image as XlsxImage
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
except ImportError as exc:
    missing = str(exc).split("No module named ")[-1].strip("'")
    print(
        f"Не хватает библиотеки {missing}.\n"
        "Установите зависимости командой:\n"
        "  .venv/bin/pip install -r requirements.txt\n"
        "Потом запустите:\n"
        "  .venv/bin/python main.py"
    )
    sys.exit(1)

import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog, ttk

APP_TITLE = "Passport creator"
PROJECT_ROOT = Path.home() / "Documents" / "MA2_passports"
APP_DATA_DIR = Path.home() / ".passport_creator"
REMOTE_CACHE_ROOT = APP_DATA_DIR / "remote_cache" / "MA2_passports"
CONFIG_PATH = APP_DATA_DIR / "cloud_connection.json"
WEB_ENV_PATH = Path(__file__).parent / "web_app" / ".env"
ACTIVE_PROJECT_ROOT = PROJECT_ROOT
REMOTE_PROJECT_ROOT_NAME = "MA2_passports"
YANDEX_APP_ROOT = f"app:/{REMOTE_PROJECT_ROOT_NAME}"
PHOTO_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp", ".bmp", ".heic", ".heif"}
XML_EXTENSIONS = {".xml"}

BLACK = "#000000"
PANEL = "#141414"
PANEL_2 = "#202020"
YELLOW = "#ffb800"
SILVER = "#cfcfcf"
MUTED = "#9c9c9c"
RED = "#ff4545"
GREEN = "#4ee06d"


@dataclass(frozen=True)
class PresetItem:
    preset_no: str
    preset_name: str
    fixture_id: str

    @property
    def preset_label(self) -> str:
        return self.preset_name.strip() or self.preset_no.strip()

    @property
    def file_stem(self) -> str:
        return safe_filename(f"{self.preset_no}_{self.fixture_id}")


@dataclass
class PassportRow:
    preset_label: str
    fixture_id: str
    preset_no: str
    photo_path: Optional[Path] = None
    description: str = ""

    @property
    def group_key(self) -> tuple[str, str, str]:
        return (self.preset_label, self.fixture_id, self.preset_no)

    @property
    def file_stem(self) -> str:
        return safe_filename(f"{self.preset_no}_{self.fixture_id}")


@dataclass
class PartituraRow:
    number: str
    name: str
    fade: str
    downfade: str
    delay: str
    trigger: str
    trigger_time: str
    info: str
    command: str
    xml_key: str = ""
    original_name: str = ""
    original_info: str = ""

    def value(self, field_id: str) -> str:
        return {
            "number": self.number,
            "name": self.name,
            "fade": self.fade,
            "downfade": self.downfade,
            "delay": self.delay,
            "trigger": self.trigger,
            "trigger_time": self.trigger_time,
            "info": self.info,
            "command": self.command,
        }.get(field_id, "")


@dataclass
class PartituraField:
    field_id: str
    title: str
    enabled: bool


@dataclass(frozen=True)
class Project:
    title: str
    directory: Path
    xml_path: Path


@dataclass
class SftpConfig:
    provider: str = "sftp"
    host: str = ""
    port: int = 22
    username: str = ""
    password: str = ""
    remote_dir: str = REMOTE_PROJECT_ROOT_NAME


@dataclass
class YandexDiskConfig:
    access_token: str = ""
    refresh_token: str = ""


@dataclass(frozen=True)
class RemoteFile:
    remote_path: str
    relative_path: Path
    size: int = 0


PARTITURA_DEFAULT_FIELDS = [
    PartituraField("number", "Номер", True),
    PartituraField("name", "Реплика", True),
    PartituraField("trigger", "Trigger", False),
    PartituraField("trigger_time", "Trigger time", False),
    PartituraField("fade", "Fade", True),
    PartituraField("downfade", "Downfade", False),
    PartituraField("delay", "Delay", False),
    PartituraField("info", "Инфо", True),
    PartituraField("command", "Command", False),
]


def local_name(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]


def child_by_name(parent: ET.Element, name: str) -> Optional[ET.Element]:
    for child in parent:
        if local_name(child.tag) == name:
            return child
    return None


def children_by_name(parent: ET.Element, name: str) -> list[ET.Element]:
    return [child for child in parent if local_name(child.tag) == name]


def first_non_empty(*values: Optional[str]) -> str:
    for value in values:
        text = (value or "").strip()
        if text:
            return text
    return ""


def natural_key(value: str) -> tuple[int, int | str]:
    return (0, int(value)) if value.isdigit() else (1, value)


def safe_filename(value: str) -> str:
    value = re.sub(r"[^0-9A-Za-zА-Яа-яЁё_. -]+", "_", value, flags=re.UNICODE)
    value = re.sub(r"\s+", " ", value).strip(" ._")
    return value or "show"


def set_active_project_root(path: Path) -> None:
    global ACTIVE_PROJECT_ROOT
    ACTIVE_PROJECT_ROOT = path


def active_project_root() -> Path:
    return ACTIVE_PROJECT_ROOT


def project_dir_for_title(title: str) -> Path:
    return active_project_root() / safe_filename(title)


def display_title(value: str) -> str:
    text = value[:-9] if value.endswith("_passport") else value
    return text.replace("_", " ")


def project_title_from_dir(path: Path) -> str:
    return display_title(path.name)


def ensure_project_root() -> None:
    active_project_root().mkdir(parents=True, exist_ok=True)


def project_xml_path(project_dir: Path) -> Optional[Path]:
    xmls = sorted(path for path in project_dir.iterdir() if path.suffix.lower() in XML_EXTENSIONS and path.is_file())
    return xmls[0] if xmls else None


def require_project_xml(project_dir: Path) -> Path:
    xml_path = project_xml_path(project_dir)
    if xml_path is None:
        raise RuntimeError(f"В папке проекта нет XML: {project_dir}")
    return xml_path


def list_projects() -> list[Project]:
    ensure_project_root()
    projects: list[Project] = []
    for directory in sorted([path for path in active_project_root().iterdir() if path.is_dir()], key=lambda p: p.name.lower()):
        xml_path = project_xml_path(directory)
        if xml_path:
            projects.append(Project(display_title(directory.name), directory, xml_path))
    return projects


def copy_xml_to_project(source: Path, title: str) -> Project:
    ensure_project_root()
    project_dir = project_dir_for_title(title)
    project_dir.mkdir(parents=True, exist_ok=True)
    target = project_dir / f"{safe_filename(title)}.xml"
    if source.resolve() != target.resolve():
        shutil.copy2(source, target)
    return Project(display_title(title), project_dir, target)


def preset_xlsx_path(project_dir: Path, title: str) -> Path:
    return project_dir / f"{safe_filename(title)}_пресеты.xlsx"


def preset_pdf_path(project_dir: Path, title: str) -> Path:
    return project_dir / f"{safe_filename(title)}_пресеты.pdf"


def partitura_xlsx_path(project_dir: Path, title: str) -> Path:
    return project_dir / f"{safe_filename(title)}_партитура.xlsx"


def partitura_pdf_path(project_dir: Path, title: str) -> Path:
    return project_dir / f"{safe_filename(title)}_партитура.pdf"


def partitura_new_xml_path(project_dir: Path, title: str) -> Path:
    return project_dir / f"{safe_filename(title)}_new.xml"


def partitura_state_path(project_dir: Path) -> Path:
    return project_dir / "partitura_state.json"


def photos_dir(project_dir: Path) -> Path:
    path = project_dir / "photos"
    path.mkdir(parents=True, exist_ok=True)
    return path


def find_existing_presets_xlsx(project_dir: Path) -> Optional[Path]:
    candidates = sorted(path for path in project_dir.glob("*_пресеты.xlsx") if not path.name.startswith("._"))
    if candidates:
        return candidates[0]
    legacy = project_dir / "passport.xlsx"
    return legacy if legacy.exists() else None


def load_dotenv_values(path: Path = WEB_ENV_PATH) -> dict[str, str]:
    values: dict[str, str] = {}
    try:
        for raw_line in path.read_text(encoding="utf-8").splitlines():
            line = raw_line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, value = line.split("=", 1)
            values[key.strip()] = value.strip().strip("\"'")
    except OSError:
        pass
    return values


def yandex_oauth_credentials() -> tuple[str, str]:
    env = load_dotenv_values()
    return env.get("YANDEX_CLIENT_ID", ""), env.get("YANDEX_CLIENT_SECRET", "")


def load_sftp_config() -> SftpConfig:
    try:
        data = json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return SftpConfig()
    raw_host = str(data.get("host", data.get("url", ""))).strip()
    raw_host = raw_host.removeprefix("sftp://").removeprefix("ssh://")
    raw_host = re.sub(r"^https?://", "", raw_host).split("/", 1)[0]
    try:
        port = int(data.get("port", 22))
    except (TypeError, ValueError):
        port = 22
    return SftpConfig(
        provider=str(data.get("provider", "sftp")),
        host=raw_host,
        port=port,
        username=str(data.get("username", "")),
        password=str(data.get("password", "")),
        remote_dir=str(data.get("remote_dir", REMOTE_PROJECT_ROOT_NAME)) or REMOTE_PROJECT_ROOT_NAME,
    )


def save_sftp_config(config: SftpConfig) -> None:
    save_cloud_config(config, load_yandex_config())


def load_yandex_config() -> YandexDiskConfig:
    try:
        data = json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return YandexDiskConfig()
    return YandexDiskConfig(
        access_token=str(data.get("yandex_access_token", "")),
        refresh_token=str(data.get("yandex_refresh_token", "")),
    )


def save_cloud_config(sftp: SftpConfig, yandex: YandexDiskConfig) -> None:
    APP_DATA_DIR.mkdir(parents=True, exist_ok=True)
    CONFIG_PATH.write_text(
        json.dumps(
            {
                "provider": sftp.provider,
                "host": sftp.host,
                "port": sftp.port,
                "username": sftp.username,
                "password": sftp.password,
                "remote_dir": sftp.remote_dir,
                "yandex_access_token": yandex.access_token,
                "yandex_refresh_token": yandex.refresh_token,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )


def local_project_files(local_dir: Path) -> list[Path]:
    return sorted(
        [path for path in local_dir.rglob("*") if path.is_file() and not path.name.endswith(".download")],
        key=lambda path: (0 if path.suffix.lower() in XML_EXTENSIONS else 1, str(path.relative_to(local_dir)).lower()),
    )


def sftp_join(*parts: str) -> str:
    cleaned = []
    absolute = False
    for part in parts:
        if not part:
            continue
        text = str(part).replace("\\", "/")
        if text.startswith("/") and not cleaned:
            absolute = True
        cleaned.extend(piece for piece in text.split("/") if piece)
    prefix = "/" if absolute else ""
    return prefix + "/".join(cleaned)


def sftp_parent(path: str) -> str:
    clean = path.rstrip("/")
    if "/" not in clean:
        return ""
    return clean.rsplit("/", 1)[0]


def sftp_basename(path: str) -> str:
    return path.rstrip("/").rsplit("/", 1)[-1]


def sftp_exists(sftp: paramiko.SFTPClient, path: str) -> bool:
    try:
        sftp.stat(path)
        return True
    except FileNotFoundError:
        return False
    except OSError:
        return False


def ensure_sftp_dir(sftp: paramiko.SFTPClient, path: str) -> None:
    current = "/" if path.startswith("/") else ""
    for part in [piece for piece in path.split("/") if piece]:
        current = sftp_join(current, part)
        try:
            sftp.mkdir(current)
        except OSError:
            pass


def sftp_is_dir(sftp: paramiko.SFTPClient, path: str) -> bool:
    try:
        return stat.S_ISDIR(sftp.stat(path).st_mode)
    except OSError:
        return False


def remove_sftp_path(sftp: paramiko.SFTPClient, path: str) -> None:
    if not sftp_exists(sftp, path):
        return
    if sftp_is_dir(sftp, path):
        for item in sftp.listdir_attr(path):
            remove_sftp_path(sftp, sftp_join(path, item.filename))
        sftp.rmdir(path)
    else:
        sftp.remove(path)


def sftp_project_names(sftp: paramiko.SFTPClient, remote_root: str) -> list[str]:
    ensure_sftp_dir(sftp, remote_root)
    names: list[str] = []
    for item in sftp.listdir_attr(remote_root):
        path = sftp_join(remote_root, item.filename)
        if stat.S_ISDIR(item.st_mode):
            try:
                if any(name.lower().endswith(".xml") for name in sftp.listdir(path)):
                    names.append(item.filename)
            except OSError:
                pass
    return sorted(names, key=str.lower)


def download_sftp_project_index(sftp: paramiko.SFTPClient, remote_root: str, local_root: Path) -> None:
    if local_root.exists():
        shutil.rmtree(local_root)
    local_root.mkdir(parents=True, exist_ok=True)
    for project_name in sftp_project_names(sftp, remote_root):
        remote_project = sftp_join(remote_root, project_name)
        local_project = local_root / project_name
        local_project.mkdir(parents=True, exist_ok=True)
        for item in sftp.listdir_attr(remote_project):
            if stat.S_ISDIR(item.st_mode):
                continue
            sftp.get(sftp_join(remote_project, item.filename), str(local_project / item.filename))


def collect_sftp_files(sftp: paramiko.SFTPClient, remote_dir: str, base_relative: Path = Path()) -> list[RemoteFile]:
    files: list[RemoteFile] = []
    for item in sftp.listdir_attr(remote_dir):
        remote_path = sftp_join(remote_dir, item.filename)
        relative = base_relative / item.filename
        if stat.S_ISDIR(item.st_mode):
            files.extend(collect_sftp_files(sftp, remote_path, relative))
        else:
            files.append(RemoteFile(remote_path, relative, int(item.st_size or 0)))
    return files


def download_sftp_project_atomic(
    sftp: paramiko.SFTPClient,
    remote_project: str,
    local_project: Path,
    progress: Optional[Callable[[int, int, str], None]] = None,
) -> None:
    if progress:
        progress(0, 0, "собираю список файлов")
    files = collect_sftp_files(sftp, remote_project)
    temp_dir = local_project.parent / f".{local_project.name}.download"
    backup_dir = local_project.parent / f".{local_project.name}.old"
    if temp_dir.exists():
        shutil.rmtree(temp_dir)
    temp_dir.mkdir(parents=True, exist_ok=True)
    try:
        total = len(files)
        for index, remote_file in enumerate(files, start=1):
            if progress:
                progress(index, total, str(remote_file.relative_path))
            local_path = temp_dir / remote_file.relative_path
            local_path.parent.mkdir(parents=True, exist_ok=True)
            sftp.get(remote_file.remote_path, str(local_path))
        if not project_xml_path(temp_dir):
            raise RuntimeError("В скачанном проекте нет XML.")
        if backup_dir.exists():
            shutil.rmtree(backup_dir)
        if local_project.exists():
            local_project.replace(backup_dir)
        temp_dir.replace(local_project)
        if backup_dir.exists():
            shutil.rmtree(backup_dir)
    except Exception:
        if temp_dir.exists():
            shutil.rmtree(temp_dir)
        raise


def upload_sftp_project_atomic(
    sftp: paramiko.SFTPClient,
    local_project: Path,
    remote_project: str,
    progress: Optional[Callable[[int, int, str], None]] = None,
) -> None:
    require_project_xml(local_project)
    remote_root = sftp_parent(remote_project)
    ensure_sftp_dir(sftp, remote_root)
    temp_remote = remote_project.rstrip("/") + ".upload"
    remove_sftp_path(sftp, temp_remote)
    ensure_sftp_dir(sftp, temp_remote)
    try:
        files = local_project_files(local_project)
        total = len(files)
        for index, local_path in enumerate(files, start=1):
            relative = local_path.relative_to(local_project)
            if progress:
                progress(index, total, str(relative))
            remote_parent = temp_remote
            for part in relative.parts[:-1]:
                remote_parent = sftp_join(remote_parent, part)
                ensure_sftp_dir(sftp, remote_parent)
            sftp.put(str(local_path), sftp_join(remote_parent, relative.name))
        remove_sftp_path(sftp, remote_project)
        try:
            sftp.rename(temp_remote, remote_project)
        except OSError:
            ensure_sftp_dir(sftp, sftp_parent(remote_project))
            sftp.rename(temp_remote, remote_project)
    except Exception:
        remove_sftp_path(sftp, temp_remote)
        raise


def copy_project_dir(source: Path, target_root: Path, replace: bool) -> Path:
    target = target_root / source.name
    if target.exists():
        if not replace:
            raise FileExistsError(target)
        shutil.rmtree(target)
    target_root.mkdir(parents=True, exist_ok=True)
    shutil.copytree(source, target)
    return target


def mirror_project_to_remote_cache(project_dir: Path) -> None:
    target = REMOTE_CACHE_ROOT / project_dir.name
    try:
        if target.resolve() == project_dir.resolve():
            return
    except OSError:
        pass
    if target.exists():
        shutil.rmtree(target)
    REMOTE_CACHE_ROOT.mkdir(parents=True, exist_ok=True)
    shutil.copytree(project_dir, target)


def https_context() -> ssl.SSLContext:
    return ssl.create_default_context(cafile=certifi.where())


def https_urlopen(request_or_url, *, timeout: int):
    return urllib.request.urlopen(request_or_url, timeout=timeout, context=https_context())


def yandex_request(
    method: str,
    endpoint: str,
    token: str,
    *,
    params: Optional[dict[str, str]] = None,
    data: bytes | None = None,
    headers: Optional[dict[str, str]] = None,
    timeout: int = 60,
) -> tuple[int, bytes]:
    query = urllib.parse.urlencode(params or {})
    url = f"https://cloud-api.yandex.net/v1/disk/{endpoint}"
    if query:
        url = f"{url}?{query}"
    request_headers = {"Authorization": f"OAuth {token}"}
    if headers:
        request_headers.update(headers)
    request = urllib.request.Request(url, data=data, headers=request_headers, method=method)
    try:
        with https_urlopen(request, timeout=timeout) as response:
            return int(response.status), response.read()
    except urllib.error.HTTPError as exc:
        body = exc.read()
        return int(exc.code), body


def yandex_json(method: str, endpoint: str, token: str, *, params: Optional[dict[str, str]] = None, timeout: int = 60) -> dict:
    status, body = yandex_request(method, endpoint, token, params=params, timeout=timeout)
    if status not in {200, 201, 202, 204, 409}:
        detail = body.decode("utf-8", "ignore")[:500]
        raise RuntimeError(f"Яндекс.Диск HTTP {status}: {detail}")
    if not body:
        return {}
    return json.loads(body.decode("utf-8"))


def yandex_upload_url(token: str, disk_path: str) -> str:
    payload = yandex_json("GET", "resources/upload", token, params={"path": disk_path, "overwrite": "true"})
    href = payload.get("href")
    if not href:
        raise RuntimeError("Яндекс.Диск не вернул ссылку загрузки.")
    return href


def yandex_download_url(token: str, disk_path: str) -> str:
    payload = yandex_json("GET", "resources/download", token, params={"path": disk_path})
    href = payload.get("href")
    if not href:
        raise RuntimeError("Яндекс.Диск не вернул ссылку скачивания.")
    return href


def yandex_put_bytes(url: str, content: bytes) -> None:
    request = urllib.request.Request(url, data=content, method="PUT")
    with https_urlopen(request, timeout=120) as response:
        if int(response.status) not in {200, 201, 202}:
            raise RuntimeError(f"Загрузка файла вернула HTTP {response.status}")


def yandex_get_bytes(url: str) -> bytes:
    with https_urlopen(url, timeout=120) as response:
        return response.read()


def yandex_ensure_dir(token: str, disk_path: str) -> None:
    current = "app:"
    for part in [piece for piece in disk_path.removeprefix("app:/").split("/") if piece]:
        current = f"{current}/{part}"
        status, body = yandex_request("PUT", "resources", token, params={"path": current}, timeout=30)
        if status not in {201, 409}:
            detail = body.decode("utf-8", "ignore")[:500]
            raise RuntimeError(f"Не удалось создать папку {current}: HTTP {status} {detail}")


def yandex_project_path(project_name: str) -> str:
    return f"{YANDEX_APP_ROOT}/{project_name}"


def yandex_relative_path(root: Path, file: Path) -> str:
    return root.relative_to(root) if False else str(file.relative_to(root)).replace("\\", "/")


def yandex_upload_project(
    token: str,
    local_project: Path,
    progress: Optional[Callable[[int, int, str], None]] = None,
) -> None:
    require_project_xml(local_project)
    remote_project = yandex_project_path(local_project.name)
    yandex_ensure_dir(token, remote_project)
    files = local_project_files(local_project)
    total = len(files)
    for index, file in enumerate(files, start=1):
        relative = str(file.relative_to(local_project)).replace("\\", "/")
        if progress:
            progress(index, total, relative)
        parent = f"{remote_project}/{relative.rsplit('/', 1)[0]}" if "/" in relative else remote_project
        yandex_ensure_dir(token, parent)
        href = yandex_upload_url(token, f"{remote_project}/{relative}")
        yandex_put_bytes(href, file.read_bytes())
    mirror_project_to_remote_cache(local_project)


def yandex_list_dir(token: str, disk_path: str, limit: int = 1000) -> list[dict]:
    payload = yandex_json("GET", "resources", token, params={"path": disk_path, "limit": str(limit)}, timeout=60)
    return payload.get("_embedded", {}).get("items", [])


def yandex_project_names(token: str) -> list[str]:
    yandex_ensure_dir(token, YANDEX_APP_ROOT)
    names = []
    for item in yandex_list_dir(token, YANDEX_APP_ROOT):
        if item.get("type") != "dir":
            continue
        try:
            entries = yandex_list_dir(token, item.get("path") or yandex_project_path(item["name"]))
        except Exception:
            continue
        if any(entry.get("type") == "file" and str(entry.get("name", "")).lower().endswith(".xml") for entry in entries):
            names.append(item.get("name", ""))
    return sorted([name for name in names if name], key=str.lower)


def yandex_collect_files(token: str, disk_path: str, relative: Path = Path()) -> list[RemoteFile]:
    files: list[RemoteFile] = []
    for item in yandex_list_dir(token, disk_path):
        item_name = item.get("name", "")
        item_path = item.get("path") or f"{disk_path}/{item_name}"
        item_relative = relative / item_name
        if item.get("type") == "dir":
            files.extend(yandex_collect_files(token, item_path, item_relative))
        elif item.get("type") == "file":
            files.append(RemoteFile(item_path, item_relative, int(item.get("size") or 0)))
    return files


def yandex_download_project_atomic(
    token: str,
    project_name: str,
    local_project: Path,
    progress: Optional[Callable[[int, int, str], None]] = None,
) -> None:
    remote_project = yandex_project_path(project_name)
    files = yandex_collect_files(token, remote_project)
    temp_dir = local_project.parent / f".{local_project.name}.download"
    backup_dir = local_project.parent / f".{local_project.name}.old"
    if temp_dir.exists():
        shutil.rmtree(temp_dir)
    temp_dir.mkdir(parents=True, exist_ok=True)
    try:
        total = len(files)
        for index, remote_file in enumerate(files, start=1):
            if progress:
                progress(index, total, str(remote_file.relative_path))
            target = temp_dir / remote_file.relative_path
            target.parent.mkdir(parents=True, exist_ok=True)
            target.write_bytes(yandex_get_bytes(yandex_download_url(token, remote_file.remote_path)))
        if not project_xml_path(temp_dir):
            raise RuntimeError("В скачанном проекте нет XML.")
        if backup_dir.exists():
            shutil.rmtree(backup_dir)
        if local_project.exists():
            local_project.replace(backup_dir)
        temp_dir.replace(local_project)
        if backup_dir.exists():
            shutil.rmtree(backup_dir)
    except Exception:
        if temp_dir.exists():
            shutil.rmtree(temp_dir)
        raise


def yandex_download_project_index(token: str, local_root: Path) -> None:
    if local_root.exists():
        shutil.rmtree(local_root)
    local_root.mkdir(parents=True, exist_ok=True)
    for project_name in yandex_project_names(token):
        local_project = local_root / project_name
        local_project.mkdir(parents=True, exist_ok=True)


def yandex_delete_project(token: str, project_name: str) -> None:
    status, body = yandex_request("DELETE", "resources", token, params={"path": yandex_project_path(project_name), "permanently": "true"}, timeout=60)
    if status not in {202, 204, 404}:
        raise RuntimeError(body.decode("utf-8", "ignore")[:500])


def yandex_rename_project(token: str, old_name: str, new_name: str) -> None:
    old_path = yandex_project_path(old_name)
    new_path = yandex_project_path(new_name)
    status, body = yandex_request("POST", "resources/move", token, params={"from": old_path, "path": new_path, "overwrite": "true"}, timeout=60)
    if status not in {201, 202}:
        raise RuntimeError(body.decode("utf-8", "ignore")[:500])


def yandex_exchange_code(code: str) -> YandexDiskConfig:
    client_id, client_secret = yandex_oauth_credentials()
    if not client_id or not client_secret:
        raise RuntimeError(f"Не найдены YANDEX_CLIENT_ID/YANDEX_CLIENT_SECRET в {WEB_ENV_PATH}")
    data = urllib.parse.urlencode(
        {
            "grant_type": "authorization_code",
            "code": code.strip(),
            "client_id": client_id,
            "client_secret": client_secret,
        }
    ).encode("utf-8")
    request = urllib.request.Request("https://oauth.yandex.ru/token", data=data, method="POST")
    request.add_header("Content-Type", "application/x-www-form-urlencoded")
    try:
        with https_urlopen(request, timeout=30) as response:
            payload = json.loads(response.read().decode("utf-8"))
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8", "ignore")
        raise RuntimeError(f"OAuth HTTP {exc.code}: {detail}")
    access = payload.get("access_token", "")
    if not access:
        raise RuntimeError("Yandex не вернул access_token.")
    return YandexDiskConfig(access, payload.get("refresh_token", ""))


def parse_grandma2_presets(xml_path: Path) -> list[PresetItem]:
    tree = ET.parse(xml_path)
    root = tree.getroot()
    ordered: dict[tuple[str, str], set[str]] = {}

    for cue_data in root.iter():
        if local_name(cue_data.tag) != "CueData":
            continue

        channel = child_by_name(cue_data, "Channel")
        preset = child_by_name(cue_data, "Preset")
        if channel is None or preset is None:
            continue

        fixture_id = channel.get("fixture_id") or channel.get("channel_id")
        preset_name = (preset.get("name") or "").strip()
        no_parts = [
            (node.text or "").strip()
            for node in children_by_name(preset, "No")
            if (node.text or "").strip()
        ]
        preset_no = ".".join(no_parts)
        if not fixture_id or not preset_name or not preset_no:
            continue

        ordered.setdefault((preset_no, preset_name), set()).add(fixture_id)

    result: list[PresetItem] = []
    for (preset_no, preset_name), fixture_ids in ordered.items():
        for fixture_id in sorted(fixture_ids, key=natural_key):
            result.append(PresetItem(preset_no, preset_name, fixture_id))
    return result


def format_cue_number(number: str, sub_number: str) -> str:
    if not sub_number or sub_number == "0":
        return number
    try:
        value = int(sub_number)
        if value % 100 == 0:
            return f"{number}.{value // 100}"
        if value % 10 == 0:
            return f"{number}.{value // 10}"
        return f"{number}.{value}"
    except ValueError:
        return f"{number}.{sub_number}"


def split_cue_number(value: str) -> tuple[str, str]:
    text = str(value or "").strip()
    if not text:
        return "0", "0"
    if "." not in text:
        return text, "0"
    number, sub = text.split(".", 1)
    sub = re.sub(r"\D", "", sub)
    if not sub:
        return number or "0", "0"
    return number or "0", str(int(sub) * 100)


def cue_xml_key(cue_number: str, order: int) -> str:
    return f"{cue_number}#{order}"


def first_index_zero_cue_part(cue: ET.Element) -> Optional[ET.Element]:
    for cue_part in cue.iter():
        if local_name(cue_part.tag) == "CuePart" and (cue_part.get("index") or "0") == "0":
            return cue_part
    return None


def iter_partitura_cues(root: ET.Element) -> list[tuple[ET.Element, str]]:
    cues: list[tuple[ET.Element, str]] = []
    order = 0
    for cue in root.iter():
        if local_name(cue.tag) != "Cue":
            continue
        number_node = child_by_name(cue, "Number")
        if number_node is None or first_index_zero_cue_part(cue) is None:
            continue
        cue_number = format_cue_number(number_node.get("number", ""), number_node.get("sub_number", "0"))
        cues.append((cue, cue_xml_key(cue_number, order)))
        order += 1
    return cues


def partitura_xml_keys(xml_path: Path) -> list[str]:
    tree = ET.parse(xml_path)
    return [key for _cue, key in iter_partitura_cues(tree.getroot())]


def parse_partitura(xml_path: Path) -> list[PartituraRow]:
    tree = ET.parse(xml_path)
    root = tree.getroot()
    rows: list[PartituraRow] = []
    order = 0

    for cue in root.iter():
        if local_name(cue.tag) != "Cue":
            continue

        number_node = child_by_name(cue, "Number")
        if number_node is None:
            continue

        cue_number = format_cue_number(number_node.get("number", ""), number_node.get("sub_number", "0"))
        trigger_node = child_by_name(cue, "Trigger")
        trigger = trigger_node.get("type") if trigger_node is not None else "Go"
        trigger_time = trigger_node.get("data_f", "") if trigger_node is not None else ""
        cue_info = ""
        cue_command = ""

        info_items = child_by_name(cue, "InfoItems")
        if info_items is not None:
            info_node = child_by_name(info_items, "Info")
            if info_node is not None and info_node.text:
                cue_info = info_node.text.strip()

        for child in cue:
            if local_name(child.tag) in {"Command", "Cmd", "CueCommand", "CLI", "CommandLine"}:
                cue_command = first_non_empty(
                    child.get("command"),
                    child.get("cmd"),
                    child.get("command_text"),
                    child.text,
                    cue_command,
                )

        for cue_part in cue.iter():
            if local_name(cue_part.tag) != "CuePart":
                continue
            if (cue_part.get("index") or "0") != "0":
                continue

            part_info = ""
            part_command = ""
            for child in cue_part.iter():
                if child is cue_part:
                    continue
                child_name = local_name(child.tag)
                if child_name == "Info" and child.text:
                    part_info = child.text.strip()
                elif child_name in {"Command", "Cmd", "CueCommand", "CLI", "CommandLine"}:
                    part_command = first_non_empty(
                        child.get("command"),
                        child.get("cmd"),
                        child.get("command_text"),
                        child.text,
                        part_command,
                    )

            rows.append(
                PartituraRow(
                    number=cue_number,
                    name=cue_part.get("name", "cue"),
                    fade=cue_part.get("basic_fade", "0"),
                    downfade=cue_part.get("basic_downfade", ""),
                    delay=cue_part.get("basic_delay", ""),
                    trigger=trigger or "Go",
                    trigger_time=trigger_time,
                    info=part_info or cue_info,
                    command=part_command or cue_command,
                    xml_key=cue_xml_key(cue_number, order),
                )
            )
            order += 1
    return rows


def qname(root: ET.Element, name: str) -> str:
    if root.tag.startswith("{"):
        namespace = root.tag.split("}", 1)[0][1:]
        return f"{{{namespace}}}{name}"
    return name


def find_cue_parent(root: ET.Element) -> ET.Element:
    for parent in root.iter():
        if any(local_name(child.tag) == "Cue" for child in parent):
            return parent
    return root


def create_empty_cue(root: ET.Element, template_cue: Optional[ET.Element]) -> ET.Element:
    cue = ET.Element(qname(root, "Cue"))
    cue.append(ET.Element(qname(root, "Number"), {"number": "0", "sub_number": "0"}))
    cue.append(ET.Element(qname(root, "CueDatas")))
    cue_part = ET.Element(qname(root, "CuePart"), {"index": "0", "name": "cue", "basic_fade": "0"})
    timing = ET.SubElement(cue_part, qname(root, "CuePartPresetTiming"))
    for _ in range(10):
        ET.SubElement(timing, qname(root, "PresetTiming"))
    cue.append(cue_part)
    if template_cue is not None and "index" in template_cue.attrib:
        cue.set("index", template_cue.get("index", "1"))
    return cue


def insert_cue_for_row(root: ET.Element, parent_map: dict[ET.Element, ET.Element], cue: ET.Element, row_index: int, rows: list[PartituraRow], cues_by_key: dict[str, ET.Element]) -> None:
    parent = find_cue_parent(root)
    insert_after: Optional[ET.Element] = None
    for previous_index in range(row_index - 1, -1, -1):
        previous_key = rows[previous_index].xml_key or f"__new_{previous_index}"
        if previous_key in cues_by_key:
            insert_after = cues_by_key[previous_key]
            break
    if insert_after is not None and parent_map.get(insert_after) is parent:
        parent.insert(list(parent).index(insert_after) + 1, cue)
    else:
        cue_children = [child for child in parent if local_name(child.tag) == "Cue"]
        if cue_children:
            parent.insert(list(parent).index(cue_children[-1]) + 1, cue)
        else:
            parent.append(cue)


def set_info(cue: ET.Element, value: str) -> None:
    text = str(value or "").strip()
    info_items = child_by_name(cue, "InfoItems")
    if not text:
        if info_items is not None:
            cue.remove(info_items)
        return
    if info_items is None:
        info_items = ET.Element(qname(cue, "InfoItems"))
        cue.append(info_items)
    info_node = child_by_name(info_items, "Info")
    if info_node is None:
        info_node = ET.Element(qname(cue, "Info"))
        info_items.append(info_node)
    info_node.text = text


def set_command(cue: ET.Element, value: str) -> None:
    text = str(value or "").strip()
    command_node = None
    for child in cue:
        if local_name(child.tag) in {"Command", "Cmd", "CueCommand", "CLI", "CommandLine"}:
            command_node = child
            break
    if not text:
        if command_node is not None:
            cue.remove(command_node)
        return
    if command_node is None:
        command_node = ET.Element(qname(cue, "Command"))
        cue.append(command_node)
    command_node.text = text


def update_cue_from_partitura_row(cue: ET.Element, row: PartituraRow) -> None:
    number_node = child_by_name(cue, "Number")
    if number_node is None:
        number_node = ET.Element(qname(cue, "Number"))
        cue.insert(0, number_node)
    number, sub_number = split_cue_number(row.number)
    number_node.set("number", number)
    number_node.set("sub_number", sub_number)

    trigger_node = child_by_name(cue, "Trigger")
    if row.trigger or row.trigger_time:
        if trigger_node is None:
            trigger_node = ET.Element(qname(cue, "Trigger"))
            cue.append(trigger_node)
        trigger_node.set("type", row.trigger or "Go")
        if row.trigger_time:
            trigger_node.set("data_f", row.trigger_time)
        elif "data_f" in trigger_node.attrib:
            del trigger_node.attrib["data_f"]
    elif trigger_node is not None:
        cue.remove(trigger_node)

    cue_part = first_index_zero_cue_part(cue)
    if cue_part is None:
        cue_part = ET.Element(qname(cue, "CuePart"), {"index": "0"})
        cue.append(cue_part)
    cue_part.set("name", row.name or "cue")
    cue_part.set("basic_fade", row.fade or "0")
    for attr, value in {"basic_downfade": row.downfade, "basic_delay": row.delay}.items():
        if value:
            cue_part.set(attr, value)
        elif attr in cue_part.attrib:
            del cue_part.attrib[attr]

    set_info(cue, row.info)
    set_command(cue, row.command)


def renumber_cue_indices(root: ET.Element) -> None:
    for parent in root.iter():
        cue_index = 1
        for child in parent:
            if local_name(child.tag) != "Cue":
                continue
            if child.get("{http://www.w3.org/2001/XMLSchema-instance}nil") == "true":
                continue
            if child_by_name(child, "Number") is None:
                continue
            child.set("index", str(cue_index))
            cue_index += 1


def write_xml_preserving_preamble(tree: ET.ElementTree, source_path: Path, output_path: Path) -> None:
    buffer = io.BytesIO()
    tree.write(buffer, encoding="utf-8", xml_declaration=True)
    content = buffer.getvalue()
    preamble = []
    try:
        for raw_line in source_path.read_text(encoding="utf-8-sig").splitlines()[:10]:
            stripped = raw_line.strip()
            if stripped.startswith("<?xml-stylesheet"):
                preamble.append(stripped)
    except OSError:
        preamble = []
    if not preamble:
        output_path.write_bytes(content)
        return
    lines = content.splitlines(keepends=True)
    insert = "".join(f"{line}\n" for line in preamble).encode("utf-8")
    if lines:
        output_path.write_bytes(lines[0] + insert + b"".join(lines[1:]))
    else:
        output_path.write_bytes(insert + content)


def save_partitura_to_show_xml(xml_path: Path, rows: list[PartituraRow], output_path: Path) -> None:
    parser = ET.XMLParser(encoding="utf-8")
    tree = ET.parse(xml_path, parser=parser)
    root = tree.getroot()
    namespace = root.tag.split("}", 1)[0][1:] if root.tag.startswith("{") else ""
    if namespace:
        ET.register_namespace("", namespace)
    if root.attrib.get("{http://www.w3.org/2001/XMLSchema-instance}schemaLocation") is not None:
        ET.register_namespace("xsi", "http://www.w3.org/2001/XMLSchema-instance")
    parent_map = {child: parent for parent in root.iter() for child in parent}
    cues_by_key = {key: cue for cue, key in iter_partitura_cues(root)}
    all_cues_by_key = dict(cues_by_key)
    used_keys: set[str] = set()
    template_cue = next((cue for cue, _key in iter_partitura_cues(root)), None)
    for row_index, row in enumerate(rows):
        key = row.xml_key or f"__new_{row_index}"
        cue = cues_by_key.get(row.xml_key or "")
        if cue is None:
            cue = create_empty_cue(root, template_cue)
            insert_cue_for_row(root, parent_map, cue, row_index, rows, all_cues_by_key)
            parent_map[cue] = find_cue_parent(root)
            all_cues_by_key[key] = cue
        else:
            used_keys.add(row.xml_key or "")
        update_cue_from_partitura_row(cue, row)
    for key, cue in cues_by_key.items():
        if key in used_keys:
            continue
        parent = parent_map.get(cue)
        if parent is not None:
            parent.remove(cue)
    renumber_cue_indices(root)
    ET.indent(tree, space="\t")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    write_xml_preserving_preamble(tree, xml_path, output_path)


def find_photo_for_stem(stem: str, photo_dir: Path) -> Optional[Path]:
    for extension in [".jpg", ".jpeg", ".png", ".webp", ".bmp"]:
        exact = photo_dir / f"{stem}{extension}"
        if exact.exists():
            return exact
    matches = sorted(
        path
        for path in photo_dir.glob(f"{stem}_*")
        if path.suffix.lower() in PHOTO_EXTENSIONS and path.is_file()
    )
    return matches[0] if matches else None


def read_passport_rows_from_sheet(sheet, item_by_group: dict[tuple[str, str], PresetItem]) -> list[PassportRow]:
    header_row = 2 if sheet.max_row >= 2 and sheet.cell(2, 1).value == "Пресет" else 1
    if sheet.cell(header_row, 1).value != "Пресет":
        return []

    rows: list[PassportRow] = []
    last_preset = ""
    last_fixture = ""
    for values in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        preset, fixture, _photo, description = (list(values) + [None, None, None, None])[:4]
        preset_text = str(preset).strip() if preset is not None else last_preset
        fixture_text = str(fixture).strip() if fixture is not None else last_fixture
        description_text = "" if description is None else str(description)
        if not preset_text and not fixture_text and not description_text.strip():
            continue
        last_preset = preset_text
        last_fixture = fixture_text
        item = item_by_group.get((preset_text, fixture_text))
        preset_no = item.preset_no if item else ""
        rows.append(PassportRow(preset_text, fixture_text, preset_no, None, description_text))
    return rows


def partitura_row_to_dict(row: PartituraRow) -> dict:
    data = {
        "number": row.number,
        "name": row.name,
        "fade": row.fade,
        "downfade": row.downfade,
        "delay": row.delay,
        "trigger": row.trigger,
        "trigger_time": row.trigger_time,
        "info": row.info,
        "command": row.command,
        "_xml_key": row.xml_key,
    }
    if row.original_name or row.original_info:
        data["_translation_original"] = {"name": row.original_name, "info": row.original_info}
    return data


def partitura_row_from_dict(data: dict) -> PartituraRow:
    original = data.get("_translation_original") or {}
    return PartituraRow(
        number=str(data.get("number", "")),
        name=str(data.get("name", "")),
        fade=str(data.get("fade", "")),
        downfade=str(data.get("downfade", "")),
        delay=str(data.get("delay", "")),
        trigger=str(data.get("trigger", "")),
        trigger_time=str(data.get("trigger_time", "")),
        info=str(data.get("info", "")),
        command=str(data.get("command", "")),
        xml_key=str(data.get("_xml_key", "")),
        original_name=str(original.get("name", "")),
        original_info=str(original.get("info", "")),
    )


def write_partitura_state(project_dir: Path, rows: list[PartituraRow], translated: bool) -> None:
    payload = {"translated": translated, "rows": [partitura_row_to_dict(row) for row in rows]}
    partitura_state_path(project_dir).write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def read_partitura_state(project_dir: Path) -> tuple[list[PartituraRow], bool]:
    path = partitura_state_path(project_dir)
    if not path.exists():
        return [], False
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return [], False
    return [partitura_row_from_dict(item) for item in payload.get("rows", []) if isinstance(item, dict)], bool(payload.get("translated"))


def load_partitura_rows(project: Project) -> tuple[list[PartituraRow], bool]:
    state_rows, translated = read_partitura_state(project.directory)
    if state_rows:
        parsed_by_key = {row.xml_key: row for row in parse_partitura(project.xml_path) if row.xml_key}
        for row in state_rows:
            if not row.xml_key:
                continue
            parsed = parsed_by_key.get(row.xml_key)
            if parsed and not row.number:
                row.number = parsed.number
        return state_rows, translated
    rows = parse_partitura(project.xml_path)
    write_partitura_state(project.directory, rows, False)
    return rows, False


def rows_for_original_partitura(rows: list[PartituraRow], translated: bool) -> list[PartituraRow]:
    if not translated:
        return rows
    result: list[PartituraRow] = []
    for row in rows:
        copy_row = PartituraRow(
            row.number,
            row.original_name or row.name,
            row.fade,
            row.downfade,
            row.delay,
            row.trigger,
            row.trigger_time,
            row.original_info or row.info,
            row.command,
            row.xml_key,
            row.original_name,
            row.original_info,
        )
        result.append(copy_row)
    return result


def load_passport_rows(items: list[PresetItem], project_dir: Path, title: str) -> tuple[list[PassportRow], Optional[str]]:
    item_by_group = {(item.preset_label, item.fixture_id): item for item in items}
    photo_dir = photos_dir(project_dir)
    xlsx_path = find_existing_presets_xlsx(project_dir)
    warning = None
    rows: list[PassportRow] = []

    if xlsx_path is not None:
        try:
            workbook = load_workbook(xlsx_path, data_only=True)
            rows = read_passport_rows_from_sheet(workbook.active, item_by_group)
        except (BadZipFile, OSError, ValueError) as exc:
            warning = f"Не удалось прочитать таблицу {xlsx_path.name}: {exc}"

    if not rows:
        rows = [PassportRow(item.preset_label, item.fixture_id, item.preset_no) for item in items]

    existing_groups = {(row.preset_label, row.fixture_id) for row in rows}
    for item in items:
        if (item.preset_label, item.fixture_id) not in existing_groups:
            rows.append(PassportRow(item.preset_label, item.fixture_id, item.preset_no))

    counters: dict[str, int] = {}
    for row in rows:
        if not row.preset_no:
            item = item_by_group.get((row.preset_label, row.fixture_id))
            row.preset_no = item.preset_no if item else ""
        stem = row.file_stem
        counters[stem] = counters.get(stem, 0) + 1
        stems = [stem]
        legacy_stem = safe_filename(f"{row.preset_label}_{row.fixture_id}")
        if legacy_stem not in stems:
            stems.append(legacy_stem)
        if counters[stem] == 1:
            row.photo_path = next((photo for candidate in stems if (photo := find_photo_for_stem(candidate, photo_dir))), None)
        else:
            row.photo_path = next((photo for candidate in stems if (photo := find_photo_for_stem(f"{candidate}_{counters[stem]}", photo_dir))), None)
    return rows, warning


def source_from_text(value: str):
    value = value.strip().split()[0] if value.strip() else "0"
    return int(value) if value.isdigit() else value


def camera_backends_for_source(source) -> list[int]:
    if isinstance(source, int) and sys.platform == "darwin":
        return [cv2.CAP_AVFOUNDATION, cv2.CAP_ANY]
    return [cv2.CAP_ANY]


def open_capture(source):
    for backend in camera_backends_for_source(source):
        capture = cv2.VideoCapture(source, backend)
        if not capture.isOpened():
            capture.release()
            continue
        capture.set(cv2.CAP_PROP_BUFFERSIZE, 1)
        return capture, backend
    return None, None


def camera_worker(source_value, frame_queue, stop_event) -> None:
    capture, backend = open_capture(source_value)
    if capture is None:
        frame_queue.put(("error", "Камера не открылась"))
        return

    backend_name = "AVFoundation" if backend == cv2.CAP_AVFOUNDATION else "default"
    deadline = time.time() + 3
    connected = False
    try:
        while not stop_event.is_set():
            ok, frame = capture.read()
            if not ok or frame is None:
                if not connected and time.time() >= deadline:
                    frame_queue.put(("error", "Камера открылась, но не дала кадр"))
                    return
                time.sleep(0.04)
                continue

            ok, encoded = cv2.imencode(".jpg", frame, [int(cv2.IMWRITE_JPEG_QUALITY), 92])
            if not ok:
                continue
            if not connected:
                frame_queue.put(("connected", backend_name))
                connected = True
            try:
                frame_queue.put_nowait(("frame", encoded.tobytes(), time.time()))
            except queue.Full:
                pass
            time.sleep(0.025)
    finally:
        capture.release()


class PlaceholderText(tk.Text):
    def __init__(self, master, placeholder: str, **kwargs):
        super().__init__(master, **kwargs)
        self.placeholder = placeholder
        self.placeholder_visible = False
        self.bind("<FocusIn>", self._hide_placeholder)
        self.bind("<FocusOut>", self._show_placeholder_if_empty)
        self._show_placeholder_if_empty()

    def _hide_placeholder(self, _event=None) -> None:
        if self.placeholder_visible:
            self.delete("1.0", "end")
            self.configure(foreground="white")
            self.placeholder_visible = False

    def _show_placeholder_if_empty(self, _event=None) -> None:
        if self.get("1.0", "end").strip():
            return
        self.placeholder_visible = True
        self.configure(foreground=MUTED)
        self.delete("1.0", "end")
        self.insert("1.0", self.placeholder)

    def real_text(self) -> str:
        if self.placeholder_visible:
            return ""
        return self.get("1.0", "end").strip()

    def set_real_text(self, value: str) -> None:
        self.placeholder_visible = False
        self.configure(foreground="white")
        self.delete("1.0", "end")
        if value:
            self.insert("1.0", value)
        self._show_placeholder_if_empty()


class PassportApp(tk.Tk):
    def __init__(self) -> None:
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("1180x780")
        self.minsize(980, 680)
        self.configure(bg=BLACK)

        self.project: Optional[Project] = None
        self.storage_mode = tk.StringVar(value="local")
        self.sftp_config = load_sftp_config()
        self.yandex_config = load_yandex_config()
        self.cloud_provider = tk.StringVar(value=self.sftp_config.provider if self.sftp_config.provider in {"sftp", "yandex_disk"} else "sftp")
        self.ssh_client: Optional[paramiko.SSHClient] = None
        self.cloud_session: Optional[paramiko.SFTPClient] = None
        self.remote_base_dir = ""
        self.remote_connected = False
        self.storage_status = tk.StringVar(value="")
        self.cloud_button_text = tk.StringVar(value="Настройки облака")
        self.transfer_status = tk.StringVar(value="")
        self.transfer_return_frame = "start"
        self.mode = "start"
        self.items: list[PresetItem] = []
        self.rows: list[PassportRow] = []
        self.current_remote_files: list[RemoteFile] = []
        self.partitura_rows: list[PartituraRow] = []
        self.partitura_translated = False
        self.partitura_fields = [PartituraField(f.field_id, f.title, f.enabled) for f in PARTITURA_DEFAULT_FIELDS]
        self.index = 0
        self.loading_description = False
        self.syncing_selection = False
        self.reviewing_photo = False
        self.captured_temp: Optional[Path] = None

        self.camera_running = False
        self.camera_opening = False
        self.camera_process: Optional[multiprocessing.Process] = None
        self.camera_queue = None
        self.camera_stop_event = None
        self.camera_last_frame_time = 0.0
        self.current_frame_bytes: Optional[bytes] = None
        self.preview_image = None

        self.frames: dict[str, tk.Frame] = {}
        self._configure_style()
        self._build_pages()
        self.update_storage_status()
        self.protocol("WM_DELETE_WINDOW", self.on_close)
        self.show_start()
        self.after(250, self.try_connect_remote_on_start)

    def _configure_style(self) -> None:
        style = ttk.Style(self)
        style.theme_use("clam")
        style.configure(".", background=BLACK, foreground="white", fieldbackground=PANEL, bordercolor=YELLOW)
        style.configure("TFrame", background=BLACK)
        style.configure("Panel.TFrame", background=PANEL)
        style.configure("TLabel", background=BLACK, foreground="white")
        style.configure("Title.TLabel", background=BLACK, foreground=YELLOW, font=("", 22, "bold"))
        style.configure("Yellow.TButton", background=BLACK, foreground=YELLOW, bordercolor=YELLOW, focusthickness=2, focuscolor=YELLOW, padding=14, font=("", 13, "bold"))
        style.map("Yellow.TButton", background=[("active", PANEL_2)], foreground=[("disabled", MUTED)])
        style.configure("Silver.TButton", background=BLACK, foreground=SILVER, bordercolor=SILVER, padding=14, font=("", 13, "bold"))
        style.map("Silver.TButton", background=[("active", PANEL_2)], foreground=[("disabled", MUTED)])
        style.configure("Danger.TButton", background=BLACK, foreground=RED, bordercolor=RED, padding=12, font=("", 13, "bold"))
        style.configure("Treeview", background=PANEL, foreground="white", fieldbackground=PANEL, rowheight=30, bordercolor=BLACK)
        style.configure("Treeview.Heading", background=BLACK, foreground=YELLOW, font=("", 11, "bold"))
        style.map("Treeview", background=[("selected", YELLOW)], foreground=[("selected", BLACK)])
        style.configure("TCheckbutton", background=BLACK, foreground="white")
        style.map("TCheckbutton", foreground=[("disabled", MUTED)])
        style.configure("TCombobox", fieldbackground=PANEL, background=PANEL, foreground="white")

    def _build_pages(self) -> None:
        container = tk.Frame(self, bg=BLACK)
        container.pack(fill="both", expand=True)
        for name in ["start", "project_source", "preset_setup", "project_list", "project_mode", "files", "workspace", "partitura", "transfer"]:
            frame = tk.Frame(container, bg=BLACK)
            frame.grid(row=0, column=0, sticky="nsew")
            self.frames[name] = frame
        container.rowconfigure(0, weight=1)
        container.columnconfigure(0, weight=1)
        self._build_start_page()
        self._build_project_source_page()
        self._build_preset_setup_page()
        self._build_project_list_page()
        self._build_project_mode_page()
        self._build_files_page()
        self._build_workspace_page()
        self._build_partitura_page()
        self._build_transfer_page()

    def show_frame(self, name: str) -> None:
        self.mode = name
        self.frames[name].tkraise()

    def _logo_widget(self, parent, large: bool = False):
        logo_path = Path(__file__).parent / "ios_app/GrandMA2Passport/GrandMA2Passport/Assets.xcassets/Logo.imageset/logoPC.png"
        if logo_path.exists():
            image = Image.open(logo_path)
            size = (240, 240) if large else (120, 120)
            image.thumbnail(size)
            photo = ImageTk.PhotoImage(image)
            label = tk.Label(parent, image=photo, bg=BLACK)
            label.image = photo
            return label
        return ttk.Label(parent, text=APP_TITLE, style="Title.TLabel")

    def _build_start_page(self) -> None:
        frame = self.frames["start"]
        frame.columnconfigure(0, weight=1)
        frame.rowconfigure(0, weight=1)
        frame.rowconfigure(1, weight=0)
        content = tk.Frame(frame, bg=BLACK)
        content.grid(row=0, column=0)
        self._logo_widget(content, large=True).pack(pady=(0, 28))
        ttk.Button(content, text="Проекты", style="Yellow.TButton", command=self.show_project_source).pack(fill="x", pady=8, ipady=8)
        ttk.Button(content, textvariable=self.cloud_button_text, style="Silver.TButton", command=self.show_connection_settings).pack(fill="x", pady=8, ipady=8)

        storage = tk.Frame(frame, bg=PANEL, highlightbackground=YELLOW, highlightthickness=1)
        storage.grid(row=1, column=0, sticky="ew", padx=120, pady=(0, 34))
        storage.columnconfigure(3, weight=1)
        tk.Label(storage, text="Хранилище", bg=PANEL, fg=SILVER, font=("", 13, "bold")).grid(row=0, column=0, padx=(18, 14), pady=14, sticky="w")
        self.local_storage_radio = tk.Radiobutton(
            storage,
            text="Локально",
            variable=self.storage_mode,
            value="local",
            command=self.on_storage_mode_changed,
            bg=PANEL,
            fg=SILVER,
            selectcolor=YELLOW,
            activebackground=PANEL,
            activeforeground=YELLOW,
        )
        self.local_storage_radio.grid(row=0, column=1, padx=10, pady=14)
        self.remote_storage_radio = tk.Radiobutton(
            storage,
            text="Облако",
            variable=self.storage_mode,
            value="remote",
            command=self.on_storage_mode_changed,
            bg=PANEL,
            fg=YELLOW,
            selectcolor=YELLOW,
            activebackground=PANEL,
            activeforeground=YELLOW,
        )
        self.remote_storage_radio.grid(row=0, column=2, padx=10, pady=14)
        self.storage_status_label = tk.Label(storage, textvariable=self.storage_status, bg=PANEL, fg=SILVER, font=("", 13, "bold"))
        self.storage_status_label.grid(row=0, column=3, sticky="w", padx=12)
        self.connection_settings_button = ttk.Button(storage, text="Настройка подключения", style="Silver.TButton", command=self.show_connection_settings)
        self.connection_settings_button.grid(row=0, column=4, padx=(10, 18), pady=10)
        storage.grid_remove()

    def _build_project_source_page(self) -> None:
        frame = self.frames["project_source"]
        frame.columnconfigure(0, weight=1)
        frame.rowconfigure(1, weight=1)
        top = tk.Frame(frame, bg=BLACK)
        top.grid(row=0, column=0, sticky="ew", padx=18, pady=18)
        ttk.Button(top, text="Назад", style="Silver.TButton", command=self.show_start).pack(side="left")
        ttk.Label(top, text="Проекты", style="Title.TLabel").pack(side="left", padx=18)

        body = tk.Frame(frame, bg=BLACK, width=560, height=320)
        body.grid(row=1, column=0)
        body.grid_propagate(False)
        body.columnconfigure(0, weight=1, minsize=560)
        ttk.Button(body, text="Устройство", style="Yellow.TButton", command=self.show_local_projects).grid(row=0, column=0, sticky="ew", pady=10, ipady=14)
        ttk.Button(body, text="Облако", style="Yellow.TButton", command=self.show_remote_projects).grid(row=1, column=0, sticky="ew", pady=10, ipady=14)
        tk.Label(body, textvariable=self.storage_status, bg=BLACK, fg=SILVER, font=("", 13, "bold"), wraplength=540, justify="center").grid(row=2, column=0, sticky="ew", pady=(22, 0))

    def _build_preset_setup_page(self) -> None:
        frame = self.frames["preset_setup"]
        frame.columnconfigure(0, weight=1)
        top = tk.Frame(frame, bg=BLACK)
        top.grid(row=0, column=0, sticky="ew", padx=18, pady=18)
        ttk.Button(top, text="Назад", style="Silver.TButton", command=self.show_start).pack(side="left")
        ttk.Label(top, text="Пресеты", style="Title.TLabel").pack(side="left", padx=18)

        body = tk.Frame(frame, bg=BLACK)
        body.grid(row=1, column=0, sticky="nsew", padx=36, pady=20)
        body.columnconfigure(0, weight=1)

        camera = tk.Frame(body, bg=PANEL)
        camera.grid(row=0, column=0, sticky="ew", pady=(0, 24))
        camera.columnconfigure(1, weight=1)
        tk.Label(camera, text="Камера", bg=PANEL, fg=SILVER, font=("", 13, "bold")).grid(row=0, column=0, padx=14, pady=14, sticky="w")
        self.camera_source_var = tk.StringVar(value="0 iPhone")
        self.camera_source = ttk.Combobox(camera, textvariable=self.camera_source_var, values=["0 iPhone", "1 FaceTime", "2", "3"], width=28)
        self.camera_source.grid(row=0, column=1, padx=8, sticky="ew")
        self.zoom_var = tk.DoubleVar(value=1.0)
        ttk.Button(camera, text="Подключить", style="Silver.TButton", command=self.connect_camera).grid(row=0, column=2, padx=8)
        ttk.Button(camera, text="Остановить", style="Silver.TButton", command=self.stop_camera).grid(row=0, column=3, padx=(0, 14))

        actions = tk.Frame(body, bg=BLACK)
        actions.grid(row=1, column=0, sticky="ew")
        actions.columnconfigure(0, weight=1)
        actions.columnconfigure(1, weight=1)
        ttk.Button(actions, text="Открыть проект", style="Yellow.TButton", command=lambda: self.show_project_list("presets")).grid(row=0, column=0, sticky="ew", padx=(0, 10), ipady=12)
        ttk.Button(actions, text="Загрузить XML", style="Yellow.TButton", command=self.create_preset_project_from_xml).grid(row=0, column=1, sticky="ew", padx=(10, 0), ipady=12)

        self.preset_setup_status = tk.StringVar(value=f"Проекты хранятся в {PROJECT_ROOT}")
        ttk.Label(body, textvariable=self.preset_setup_status).grid(row=2, column=0, sticky="w", pady=20)

    def _build_project_list_page(self) -> None:
        frame = self.frames["project_list"]
        frame.columnconfigure(0, weight=1)
        frame.rowconfigure(2, weight=1)
        top = tk.Frame(frame, bg=BLACK)
        top.grid(row=0, column=0, sticky="ew", padx=18, pady=18)
        ttk.Button(top, text="Назад", style="Silver.TButton", command=self.back_from_project_list).pack(side="left")
        self.project_list_title = tk.StringVar(value="Проекты")
        ttk.Label(top, textvariable=self.project_list_title, style="Title.TLabel").pack(side="left", padx=18)

        self.project_listbox = tk.Listbox(frame, bg=PANEL, fg="white", selectbackground=YELLOW, selectforeground=BLACK, font=("", 22), activestyle="none")
        self.project_listbox.grid(row=2, column=0, sticky="nsew", padx=36, pady=(0, 14))
        self.project_listbox.bind("<Double-Button-1>", lambda _e: self.open_selected_project())
        self.project_listbox.bind("<Return>", lambda _e: self.open_selected_project())
        self.project_listbox.bind("<Button-3>", self.show_project_menu)
        self.project_listbox.bind("<Button-2>", self.show_project_menu)

        buttons = tk.Frame(frame, bg=BLACK)
        buttons.grid(row=3, column=0, sticky="ew", padx=36, pady=(0, 24))
        buttons.columnconfigure(0, weight=1)
        self.create_project_button = ttk.Button(buttons, text="Создать проект", style="Yellow.TButton", command=self.create_preset_project_from_xml)
        self.create_project_button.grid(row=0, column=0, sticky="ew", ipady=14)

        self.project_menu = tk.Menu(self, tearoff=False)
        self.project_menu.add_command(label="Открыть", command=self.open_selected_project)
        self.project_menu.add_command(label="Переименовать", command=self.rename_selected_project)
        self.project_menu.add_command(label="Удалить", command=self.delete_selected_project)

    def _build_project_mode_page(self) -> None:
        frame = self.frames["project_mode"]
        frame.columnconfigure(0, weight=1)
        top = tk.Frame(frame, bg=BLACK)
        top.grid(row=0, column=0, sticky="ew", padx=18, pady=18)
        ttk.Button(top, text="Назад", style="Silver.TButton", command=lambda: self.show_project_list("projects")).pack(side="left")
        self.project_mode_title = tk.StringVar(value="Проект")
        ttk.Label(top, textvariable=self.project_mode_title, style="Title.TLabel").pack(side="left", padx=18)

        body = tk.Frame(frame, bg=BLACK)
        body.grid(row=1, column=0, sticky="", padx=80, pady=70)
        body.columnconfigure((0, 1), weight=1)

        self.open_presets_button = ttk.Button(body, text="Пресеты", style="Yellow.TButton", command=lambda: self.open_project_builder("presets"))
        self.open_presets_button.grid(row=0, column=0, sticky="ew", padx=18, pady=18, ipady=26)
        self.open_presets_button.bind("<Button-3>", lambda _e: self.show_kind_menu("presets"))
        self.open_presets_button.bind("<Button-2>", lambda _e: self.show_kind_menu("presets"))

        self.open_partitura_button = ttk.Button(body, text="Партитура", style="Yellow.TButton", command=lambda: self.open_project_builder("partitura"))
        self.open_partitura_button.grid(row=0, column=1, sticky="ew", padx=18, pady=18, ipady=26)
        self.open_partitura_button.bind("<Button-3>", lambda _e: self.show_kind_menu("partitura"))
        self.open_partitura_button.bind("<Button-2>", lambda _e: self.show_kind_menu("partitura"))

        self.kind_menu = tk.Menu(self, tearoff=False)

    def _build_files_page(self) -> None:
        frame = self.frames["files"]
        frame.columnconfigure(0, weight=1)
        frame.rowconfigure(1, weight=1)
        top = tk.Frame(frame, bg=BLACK)
        top.grid(row=0, column=0, sticky="ew", padx=18, pady=18)
        ttk.Button(top, text="Назад", style="Silver.TButton", command=self.back_from_files).pack(side="left")
        self.files_title = tk.StringVar(value="Файлы")
        ttk.Label(top, textvariable=self.files_title, style="Title.TLabel").pack(side="left", padx=18)

        self.files_listbox = tk.Listbox(frame, bg=PANEL, fg="white", selectbackground=YELLOW, selectforeground=BLACK, font=("", 15), activestyle="none")
        self.files_listbox.grid(row=1, column=0, sticky="nsew", padx=36, pady=(0, 14))
        self.files_listbox.bind("<Double-Button-1>", lambda _e: self.open_selected_file())
        self.files_listbox.bind("<Return>", lambda _e: self.open_selected_file())
        self.files_listbox.bind("<Button-3>", self.show_file_menu)
        self.files_listbox.bind("<Button-2>", self.show_file_menu)

        buttons = tk.Frame(frame, bg=BLACK)
        buttons.grid(row=2, column=0, sticky="ew", padx=36, pady=(0, 24))
        buttons.columnconfigure((0, 1, 2), weight=1)
        ttk.Button(buttons, text="Открыть", style="Yellow.TButton", command=self.open_selected_file).grid(row=0, column=0, sticky="ew", padx=(0, 8))
        ttk.Button(buttons, text="Показать в папке", style="Silver.TButton", command=self.reveal_selected_file).grid(row=0, column=1, sticky="ew", padx=8)
        ttk.Button(buttons, text="Удалить", style="Danger.TButton", command=self.delete_selected_file).grid(row=0, column=2, sticky="ew", padx=(8, 0))

        self.file_menu = tk.Menu(self, tearoff=False)
        self.file_menu.add_command(label="Открыть", command=self.open_selected_file)
        self.file_menu.add_command(label="Показать в папке", command=self.reveal_selected_file)
        self.file_menu.add_command(label="Удалить", command=self.delete_selected_file)

    def _build_workspace_page(self) -> None:
        frame = self.frames["workspace"]
        frame.columnconfigure(0, weight=3)
        frame.columnconfigure(1, weight=2)
        frame.rowconfigure(1, weight=1)

        top = tk.Frame(frame, bg=BLACK)
        top.grid(row=0, column=0, columnspan=2, sticky="ew", padx=14, pady=12)
        ttk.Button(top, text="Назад", style="Silver.TButton", command=self.show_project_mode).pack(side="left", padx=(0, 8))
        ttk.Button(top, text="Файлы", style="Yellow.TButton", command=lambda: self.show_files("presets")).pack(side="left", padx=8)
        self.workspace_summary = tk.StringVar(value="")
        ttk.Label(top, textvariable=self.workspace_summary).pack(side="left", padx=16)

        left = tk.Frame(frame, bg=BLACK)
        left.grid(row=1, column=0, sticky="nsew", padx=(14, 8), pady=(0, 14))
        left.columnconfigure(0, weight=1)
        left.rowconfigure(2, weight=1)

        self.current_var = tk.StringVar(value="")
        ttk.Label(left, textvariable=self.current_var, font=("", 16, "bold")).grid(row=0, column=0, sticky="ew", pady=(0, 8))

        camera = tk.Frame(left, bg=PANEL)
        camera.grid(row=1, column=0, sticky="ew", pady=(0, 10))
        camera.columnconfigure(1, weight=1)
        tk.Label(camera, text="Камера", bg=PANEL, fg=SILVER, font=("", 13, "bold")).grid(row=0, column=0, padx=14, pady=10, sticky="w")
        self.workspace_camera_source = ttk.Combobox(camera, textvariable=self.camera_source_var, values=["0 iPhone", "1 FaceTime", "2", "3"], width=28)
        self.workspace_camera_source.grid(row=0, column=1, padx=8, sticky="ew")
        ttk.Button(camera, text="Подключить", style="Silver.TButton", command=self.connect_camera).grid(row=0, column=2, padx=8)
        ttk.Button(camera, text="Остановить", style="Silver.TButton", command=self.stop_camera).grid(row=0, column=3, padx=(0, 14))

        self.video_container = tk.Frame(left, bg=PANEL)
        self.video_container.grid(row=2, column=0, sticky="nsew", pady=(0, 10))
        self.video_container.columnconfigure(0, weight=1)
        self.video_container.rowconfigure(0, weight=1)
        self.video_label = tk.Label(self.video_container, bg=PANEL, fg=SILVER, text="", font=("", 16), compound="center")
        self.video_label.grid(row=0, column=0, sticky="nsew")
        self.delete_photo_button = ttk.Button(self.video_container, text="Удалить", style="Danger.TButton", command=self.delete_current_photo)
        self.delete_photo_button.place_forget()

        description_frame = tk.Frame(left, bg=BLACK)
        description_frame.grid(row=3, column=0, sticky="nsew", pady=(0, 10))
        description_frame.columnconfigure(0, weight=1)
        description_frame.rowconfigure(0, weight=1)
        self.description_text = PlaceholderText(
            description_frame,
            "Напиши тут описание пресета.",
            height=4,
            wrap="word",
            bg=PANEL,
            fg="white",
            insertbackground=YELLOW,
            relief="flat",
            padx=10,
            pady=10,
        )
        self.description_text.grid(row=0, column=0, sticky="nsew")
        self.description_text.bind("<KeyRelease>", self.on_description_changed)

        controls = tk.Frame(left, bg=BLACK)
        controls.grid(row=4, column=0, sticky="ew")
        controls.columnconfigure((0, 1, 2, 3, 4), weight=1)
        self.photo_button = ttk.Button(controls, text="Фото", style="Yellow.TButton", command=self.take_photo)
        self.photo_button.grid(row=0, column=0, sticky="ew", padx=(0, 8))
        self.upload_photo_button = ttk.Button(controls, text="Загрузить фото", style="Yellow.TButton", command=self.upload_photo)
        self.upload_photo_button.grid(row=0, column=1, sticky="ew", padx=8)
        self.use_button = ttk.Button(controls, text="Готово", style="Yellow.TButton", command=self.use_photo)
        self.use_button.grid(row=0, column=2, sticky="ew", padx=8)
        self.retake_button = ttk.Button(controls, text="Переснять", style="Yellow.TButton", command=self.retake_photo)
        self.retake_button.grid(row=0, column=3, sticky="ew", padx=8)
        self.add_row_button = ttk.Button(controls, text="Добавить", style="Yellow.TButton", command=self.add_extra_row)
        self.add_row_button.grid(row=0, column=4, sticky="ew", padx=(8, 0))

        zoom_controls = tk.Frame(controls, bg=BLACK)
        zoom_controls.grid(row=1, column=0, columnspan=5, sticky="ew", pady=(10, 0))
        zoom_controls.columnconfigure(1, weight=1)
        tk.Label(zoom_controls, text="Зум", bg=BLACK, fg=SILVER, font=("", 11, "bold")).grid(row=0, column=0, padx=(0, 10))
        self.zoom_scale = tk.Scale(
            zoom_controls,
            from_=1.0,
            to=5.0,
            resolution=0.1,
            variable=self.zoom_var,
            orient="horizontal",
            bg=BLACK,
            fg=YELLOW,
            troughcolor=PANEL,
            highlightthickness=0,
            activebackground=YELLOW,
            command=lambda _v: self.refresh_zoom_preview(),
        )
        self.zoom_scale.grid(row=0, column=1, sticky="ew")

        right = tk.Frame(frame, bg=BLACK)
        right.grid(row=1, column=1, sticky="nsew", padx=(8, 14), pady=(0, 14))
        right.columnconfigure(0, weight=1)
        right.rowconfigure(1, weight=1)
        self.progress_var = tk.StringVar(value="0 / 0")
        ttk.Label(right, textvariable=self.progress_var, font=("", 13, "bold")).grid(row=0, column=0, sticky="ew", pady=(0, 8))
        columns = ("photo", "preset", "fixture", "description")
        self.table = ttk.Treeview(right, columns=columns, show="headings", height=22)
        self.table.heading("photo", text="Фото")
        self.table.heading("preset", text="Пресет")
        self.table.heading("fixture", text="Прибор")
        self.table.heading("description", text="Описание")
        self.table.column("photo", width=54, anchor="center")
        self.table.column("preset", width=230)
        self.table.column("fixture", width=80, anchor="center")
        self.table.column("description", width=160)
        self.table.grid(row=1, column=0, sticky="nsew")
        scroll = ttk.Scrollbar(right, orient="vertical", command=self.table.yview)
        scroll.grid(row=1, column=1, sticky="ns")
        self.table.configure(yscrollcommand=scroll.set)
        self.table.bind("<<TreeviewSelect>>", self.on_table_select)
        self.table.bind("<Button-3>", self.show_table_menu)
        self.table.bind("<Button-2>", self.show_table_menu)
        self.table_menu = tk.Menu(self, tearoff=False)
        self.table_menu.add_command(label="Удалить запись", command=self.delete_current_row)

    def _build_partitura_page(self) -> None:
        frame = self.frames["partitura"]
        frame.columnconfigure(0, weight=1)
        frame.rowconfigure(3, weight=1)
        top = tk.Frame(frame, bg=BLACK)
        top.grid(row=0, column=0, sticky="ew", padx=18, pady=18)
        ttk.Button(top, text="Назад", style="Silver.TButton", command=self.show_project_mode).pack(side="left")
        ttk.Label(top, text="Партитура", style="Title.TLabel").pack(side="left", padx=18)

        actions = tk.Frame(frame, bg=BLACK)
        actions.grid(row=1, column=0, sticky="ew", padx=36, pady=(0, 18))
        actions.columnconfigure((0, 1), weight=1)
        ttk.Button(actions, text="Открыть проект", style="Yellow.TButton", command=lambda: self.show_project_list("partitura")).grid(row=0, column=0, sticky="ew", padx=(0, 10), ipady=8)
        ttk.Button(actions, text="Загрузить XML", style="Yellow.TButton", command=self.create_partitura_project_from_xml).grid(row=0, column=1, sticky="ew", padx=(10, 0), ipady=8)
        actions.grid_remove()

        self.partitura_project_var = tk.StringVar(value="Проект не выбран")
        ttk.Label(frame, textvariable=self.partitura_project_var).grid(row=2, column=0, sticky="nw", padx=36)

        fields_frame = tk.Frame(frame, bg=BLACK)
        fields_frame.grid(row=3, column=0, sticky="nsew", padx=36, pady=18)
        fields_frame.columnconfigure(0, weight=1)
        fields_frame.rowconfigure(0, weight=1)
        self.field_listbox = tk.Listbox(fields_frame, bg=PANEL, fg="white", selectbackground=YELLOW, selectforeground=BLACK, font=("", 22), height=9, activestyle="none")
        self.field_listbox.grid(row=0, column=0, sticky="nsew")
        field_buttons = tk.Frame(fields_frame, bg=BLACK)
        field_buttons.grid(row=0, column=1, sticky="ns", padx=(12, 0))
        ttk.Button(field_buttons, text="Вкл/выкл", style="Silver.TButton", command=self.toggle_partitura_field).pack(fill="x", pady=(0, 12), ipady=8)
        ttk.Button(field_buttons, text="Вверх", style="Silver.TButton", command=lambda: self.move_partitura_field(-1)).pack(fill="x", pady=12, ipady=8)
        ttk.Button(field_buttons, text="Вниз", style="Silver.TButton", command=lambda: self.move_partitura_field(1)).pack(fill="x", pady=12, ipady=8)
        self.refresh_partitura_fields()

        bottom = tk.Frame(frame, bg=BLACK)
        bottom.grid(row=4, column=0, sticky="ew", padx=36, pady=(0, 28))
        bottom.columnconfigure((0, 1), weight=1)
        ttk.Button(bottom, text="Создать партитуру", style="Yellow.TButton", command=self.create_partitura_files).grid(row=0, column=0, sticky="ew", padx=(0, 8), ipady=10)
        ttk.Button(bottom, text="Сохранить XML", style="Yellow.TButton", command=self.save_partitura_show_xml).grid(row=0, column=1, sticky="ew", padx=(8, 0), ipady=10)

    def _build_transfer_page(self) -> None:
        frame = self.frames["transfer"]
        frame.columnconfigure(0, weight=1)
        frame.rowconfigure(0, weight=1)
        body = tk.Frame(frame, bg=BLACK, width=680, height=240)
        body.grid(row=0, column=0)
        body.grid_propagate(False)
        body.columnconfigure(0, weight=1)
        ttk.Label(body, text="Облако", style="Title.TLabel").grid(row=0, column=0, pady=(8, 26))
        tk.Label(body, textvariable=self.transfer_status, bg=BLACK, fg=SILVER, font=("", 18, "bold"), wraplength=640, justify="center").grid(row=1, column=0, sticky="ew")
        self.transfer_progress = ttk.Progressbar(body, orient="horizontal", mode="indeterminate", length=560)
        self.transfer_progress.grid(row=2, column=0, sticky="ew", padx=40, pady=(32, 0))

    def show_transfer(self, text: str) -> None:
        if self.mode != "transfer":
            self.transfer_return_frame = self.mode
        self.transfer_status.set(text)
        self.storage_status.set(text)
        self.show_frame("transfer")
        try:
            self.transfer_progress.start(12)
        except Exception:
            pass

    def hide_transfer(self, return_to_previous: bool = False) -> None:
        try:
            self.transfer_progress.stop()
        except Exception:
            pass
        if return_to_previous:
            self.show_frame(self.transfer_return_frame)

    def update_storage_status(self) -> None:
        local_selected = self.storage_mode.get() == "local"
        if hasattr(self, "local_storage_radio"):
            self.local_storage_radio.configure(fg=YELLOW if local_selected else SILVER, activeforeground=YELLOW if local_selected else SILVER)
            self.remote_storage_radio.configure(fg=YELLOW if not local_selected else SILVER, activeforeground=YELLOW if not local_selected else SILVER)
        if self.storage_mode.get() == "remote":
            self.connection_settings_button.grid()
            if self.remote_connected:
                provider = "Яндекс.Диск" if self.cloud_provider.get() == "yandex_disk" else "SFTP"
                self.storage_status.set(f"✓ {provider} подключено")
                self.storage_status_label.configure(fg=GREEN)
            else:
                self.storage_status.set("✕ нет подключения")
                self.storage_status_label.configure(fg=RED)
        else:
            self.connection_settings_button.grid_remove()
            self.storage_status.set(f"Локально: {PROJECT_ROOT}")
            self.storage_status_label.configure(fg=SILVER)
        if hasattr(self, "cloud_button_text"):
            self.cloud_button_text.set("Облако подключено" if self.remote_connected else "Настройки облака")
        if hasattr(self, "preset_setup_status"):
            self.preset_setup_status.set(f"Проекты хранятся в {active_project_root()}")

    def try_connect_remote_on_start(self) -> None:
        if self.cloud_provider.get() == "yandex_disk":
            if not self.yandex_config.access_token:
                self.update_storage_status()
                return
        elif not self.sftp_config.host or not self.sftp_config.username:
            self.update_storage_status()
            return
        self.show_transfer("Проверяю подключение к облаку...")

        def worker() -> None:
            ok, error = self.connect_remote(self.sftp_config)
            self.after(0, lambda: self.on_start_remote_connect_done(ok, error))

        threading.Thread(target=worker, daemon=True).start()

    def on_start_remote_connect_done(self, ok: bool, error: str) -> None:
        self.hide_transfer()
        self.remote_connected = ok
        self.update_storage_status()
        if not ok:
            self.storage_status.set(f"Облако не подключено: {error}")
        self.show_start()

    def on_storage_mode_changed(self) -> None:
        self.project = None
        if self.storage_mode.get() == "remote":
            if self.remote_connected:
                set_active_project_root(REMOTE_CACHE_ROOT)
            else:
                set_active_project_root(PROJECT_ROOT)
        else:
            set_active_project_root(PROJECT_ROOT)
        self.update_storage_status()

    def show_connection_settings(self) -> None:
        window = tk.Toplevel(self)
        window.title("Настройка облака")
        window.configure(bg=BLACK)
        window.transient(self)
        window.grab_set()
        window.columnconfigure(0, weight=1)

        provider = tk.StringVar(value=self.cloud_provider.get())
        tab_row = tk.Frame(window, bg=BLACK)
        tab_row.grid(row=0, column=0, sticky="ew", padx=18, pady=(18, 10))
        tab_row.columnconfigure((0, 1), weight=1)

        values = {
            "host": tk.StringVar(value=self.sftp_config.host),
            "port": tk.StringVar(value=str(self.sftp_config.port)),
            "username": tk.StringVar(value=self.sftp_config.username),
            "password": tk.StringVar(value=self.sftp_config.password),
            "remote_dir": tk.StringVar(value=self.sftp_config.remote_dir),
        }
        yandex_code = tk.StringVar(value="")
        body = tk.Frame(window, bg=BLACK)
        body.grid(row=1, column=0, sticky="nsew", padx=18, pady=8)
        body.columnconfigure(0, weight=1)

        sftp_frame = tk.Frame(body, bg=BLACK)
        sftp_frame.columnconfigure(1, weight=1)
        yandex_frame = tk.Frame(body, bg=BLACK)
        yandex_frame.columnconfigure(0, weight=1)

        labels = [("SFTP сервер", "host"), ("Порт", "port"), ("Пользователь", "username"), ("Пароль", "password"), ("Папка", "remote_dir")]
        for row, (label, key) in enumerate(labels):
            tk.Label(sftp_frame, text=label, bg=BLACK, fg=SILVER, font=("", 12, "bold")).grid(row=row, column=0, sticky="w", pady=10)
            entry = tk.Entry(sftp_frame, textvariable=values[key], bg=PANEL, fg="white", insertbackground=YELLOW, relief="flat", show="*" if key == "password" else "")
            entry.grid(row=row, column=1, sticky="ew", padx=(14, 0), pady=10, ipady=7)

        tk.Label(yandex_frame, text="1. Открой Яндекс, разреши доступ и скопируй код.", bg=BLACK, fg=SILVER, font=("", 12, "bold")).grid(row=0, column=0, sticky="w", pady=(0, 10))
        ttk.Button(yandex_frame, text="Открыть Яндекс", style="Yellow.TButton", command=lambda: self.open_yandex_oauth_url()).grid(row=1, column=0, sticky="ew", pady=(0, 14), ipady=8)
        tk.Label(yandex_frame, text="2. Вставь код сюда.", bg=BLACK, fg=SILVER, font=("", 12, "bold")).grid(row=2, column=0, sticky="w", pady=(0, 8))
        code_row = tk.Frame(yandex_frame, bg=BLACK)
        code_row.grid(row=3, column=0, sticky="ew")
        code_row.columnconfigure(0, weight=1)
        code_entry = tk.Entry(code_row, textvariable=yandex_code, bg=PANEL, fg="white", insertbackground=YELLOW, relief="flat")
        code_entry.grid(row=0, column=0, sticky="ew", ipady=8)

        def paste_code() -> None:
            try:
                yandex_code.set(self.clipboard_get().strip())
                code_entry.focus_set()
                code_entry.icursor("end")
            except tk.TclError:
                pass

        ttk.Button(code_row, text="Вставить", style="Silver.TButton", command=paste_code).grid(row=0, column=1, sticky="ew", padx=(10, 0), ipady=6)

        status = tk.StringVar(value="")
        status_label = tk.Label(window, textvariable=status, bg=BLACK, fg=SILVER)
        status_label.grid(row=2, column=0, sticky="w", padx=18, pady=(6, 0))

        def refresh_provider_view() -> None:
            for child in body.winfo_children():
                child.grid_remove()
            if provider.get() == "sftp":
                sftp_frame.grid(row=0, column=0, sticky="nsew")
            else:
                yandex_frame.grid(row=0, column=0, sticky="nsew")
            sftp_tab.configure(style="Yellow.TButton" if provider.get() == "sftp" else "Silver.TButton")
            yandex_tab.configure(style="Yellow.TButton" if provider.get() == "yandex_disk" else "Silver.TButton")

        def choose_provider(value: str) -> None:
            provider.set(value)
            status.set("")
            status_label.configure(fg=SILVER)
            refresh_provider_view()

        sftp_tab = ttk.Button(tab_row, text="SFTP", style="Yellow.TButton", command=lambda: choose_provider("sftp"))
        yandex_tab = ttk.Button(tab_row, text="Яндекс.Диск", style="Silver.TButton", command=lambda: choose_provider("yandex_disk"))
        sftp_tab.grid(row=0, column=0, sticky="ew", padx=(0, 8), ipady=8)
        yandex_tab.grid(row=0, column=1, sticky="ew", padx=(8, 0), ipady=8)
        refresh_provider_view()

        def save_and_connect() -> None:
            chosen = provider.get()
            if chosen == "yandex_disk":
                try:
                    code = yandex_code.get().strip()
                    if not code:
                        raise RuntimeError("Код не введен.")
                    config = yandex_exchange_code(code)
                except Exception as exc:
                    status.set(f"Ошибка: {exc}")
                    status_label.configure(fg=RED)
                    return
                self.yandex_config = config
                self.sftp_config.provider = "yandex_disk"
                self.cloud_provider.set("yandex_disk")
                save_cloud_config(self.sftp_config, self.yandex_config)
                self.storage_mode.set("remote")
                ok, error = self.connect_remote(self.sftp_config)
                if ok:
                    set_active_project_root(REMOTE_CACHE_ROOT)
                    self.update_storage_status()
                    window.destroy()
                else:
                    status.set(f"Ошибка: {error}")
                    status_label.configure(fg=RED)
                return

            try:
                port = int(values["port"].get().strip() or "22")
            except ValueError:
                messagebox.showerror(APP_TITLE, "Порт должен быть числом.", parent=window)
                return
            config = SftpConfig(
                provider="sftp",
                host=values["host"].get().strip(),
                port=port,
                username=values["username"].get().strip(),
                password=values["password"].get(),
                remote_dir=values["remote_dir"].get().strip() or REMOTE_PROJECT_ROOT_NAME,
            )
            if not config.host or not config.username:
                messagebox.showerror(APP_TITLE, "Заполните SFTP сервер и пользователя.", parent=window)
                return
            status.set("Подключаюсь...")
            window.update_idletasks()
            ok, error = self.connect_remote(config)
            if ok:
                save_cloud_config(config, self.yandex_config)
                self.sftp_config = config
                self.cloud_provider.set("sftp")
                self.storage_mode.set("remote")
                set_active_project_root(REMOTE_CACHE_ROOT)
                self.update_storage_status()
                window.destroy()
            else:
                status.set(f"Ошибка: {error}")
                status_label.configure(fg=RED)

        buttons = tk.Frame(window, bg=BLACK)
        buttons.grid(row=3, column=0, sticky="ew", padx=18, pady=18)
        buttons.columnconfigure((0, 1), weight=1)
        ttk.Button(buttons, text="Назад", style="Silver.TButton", command=window.destroy).grid(row=0, column=0, sticky="ew", padx=(0, 8), ipady=6)
        ttk.Button(buttons, text="Подключить и сохранить", style="Yellow.TButton", command=save_and_connect).grid(row=0, column=1, sticky="ew", padx=(8, 0), ipady=6)

    def open_yandex_oauth_url(self) -> None:
        client_id, client_secret = yandex_oauth_credentials()
        if not client_id or not client_secret:
            messagebox.showerror(APP_TITLE, f"Не найдены YANDEX_CLIENT_ID/YANDEX_CLIENT_SECRET в {WEB_ENV_PATH}")
            return
        url = "https://oauth.yandex.ru/authorize?" + urllib.parse.urlencode(
            {
                "response_type": "code",
                "client_id": client_id,
                "redirect_uri": "https://oauth.yandex.ru/verification_code",
                "scope": "cloud_api:disk.app_folder",
                "force_confirm": "yes",
            }
        )
        webbrowser.open(url)

    def connect_yandex_via_code(self, parent) -> YandexDiskConfig:
        client_id, client_secret = yandex_oauth_credentials()
        if not client_id or not client_secret:
            raise RuntimeError(f"Не найдены YANDEX_CLIENT_ID/YANDEX_CLIENT_SECRET в {WEB_ENV_PATH}")
        scope = "cloud_api:disk.app_folder"
        url = "https://oauth.yandex.ru/authorize?" + urllib.parse.urlencode(
            {
                "response_type": "code",
                "client_id": client_id,
                "redirect_uri": "https://oauth.yandex.ru/verification_code",
                "scope": scope,
                "force_confirm": "yes",
            }
        )
        webbrowser.open(url)
        code = simpledialog.askstring(APP_TITLE, "Яндекс открылся в браузере.\nСкопируйте код подтверждения и вставьте сюда:", parent=parent)
        if not code:
            raise RuntimeError("Код не введен.")
        return yandex_exchange_code(code)

    def connect_remote(self, config: SftpConfig) -> tuple[bool, str]:
        self.close_remote()
        if self.cloud_provider.get() == "yandex_disk" or config.provider == "yandex_disk":
            try:
                if not self.yandex_config.access_token:
                    return False, "Яндекс.Диск не подключен"
                yandex_ensure_dir(self.yandex_config.access_token, YANDEX_APP_ROOT)
                REMOTE_CACHE_ROOT.mkdir(parents=True, exist_ok=True)
                self.remote_base_dir = YANDEX_APP_ROOT
                self.remote_connected = True
                return True, ""
            except Exception as exc:
                self.remote_connected = False
                return False, str(exc)
        ssh: Optional[paramiko.SSHClient] = None
        try:
            ssh = paramiko.SSHClient()
            ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            ssh.connect(
                hostname=config.host,
                port=config.port,
                username=config.username,
                password=config.password,
                timeout=12,
                banner_timeout=12,
                auth_timeout=12,
            )
            session = ssh.open_sftp()
            base = config.remote_dir.strip("/") or REMOTE_PROJECT_ROOT_NAME
            ensure_sftp_dir(session, base)
            REMOTE_CACHE_ROOT.mkdir(parents=True, exist_ok=True)
        except Exception as exc:
            if ssh is not None:
                try:
                    ssh.close()
                except Exception:
                    pass
            self.remote_connected = False
            return False, str(exc)

        self.ssh_client = ssh
        self.cloud_session = session
        self.remote_base_dir = base
        self.remote_connected = True
        return True, ""

    def close_remote(self) -> None:
        if self.cloud_session is not None:
            try:
                self.cloud_session.close()
            except Exception:
                pass
        if self.ssh_client is not None:
            try:
                self.ssh_client.close()
            except Exception:
                pass
        self.ssh_client = None
        self.cloud_session = None
        self.remote_connected = False

    def ensure_remote_ready(self) -> bool:
        if self.cloud_provider.get() == "yandex_disk":
            if self.remote_connected and self.yandex_config.access_token:
                return True
            if not self.yandex_config.access_token:
                messagebox.showwarning(APP_TITLE, "Сначала подключите Яндекс.Диск.")
                return False
            ok, error = self.connect_remote(self.sftp_config)
            if ok:
                self.remote_connected = True
                if self.storage_mode.get() == "remote":
                    set_active_project_root(REMOTE_CACHE_ROOT)
                self.update_storage_status()
                return True
            self.update_storage_status()
            messagebox.showerror(APP_TITLE, f"Не удалось подключиться к Яндекс.Диску:\n{error}")
            return False
        if self.remote_connected and self.cloud_session is not None:
            return True
        if not self.sftp_config.host:
            messagebox.showwarning(APP_TITLE, "Сначала настройте подключение к облаку.")
            return False
        ok, error = self.connect_remote(self.sftp_config)
        if ok:
            self.remote_connected = True
            if self.storage_mode.get() == "remote":
                set_active_project_root(REMOTE_CACHE_ROOT)
            self.update_storage_status()
            return True
        self.update_storage_status()
        messagebox.showerror(APP_TITLE, f"Не удалось подключиться к облаку:\n{error}")
        return False

    def ensure_remote_ready_silent(self) -> bool:
        if self.cloud_provider.get() == "yandex_disk":
            if self.remote_connected and self.yandex_config.access_token:
                return True
            if not self.yandex_config.access_token:
                return False
            ok, _error = self.connect_remote(self.sftp_config)
            self.remote_connected = ok
            return ok
        if self.remote_connected and self.cloud_session is not None:
            return True
        if not self.sftp_config.host:
            return False
        ok, _error = self.connect_remote(self.sftp_config)
        self.remote_connected = ok
        return ok

    def refresh_remote_cache(self) -> bool:
        if not self.ensure_remote_ready():
            return False
        try:
            if self.cloud_provider.get() == "yandex_disk":
                yandex_download_project_index(self.yandex_config.access_token, REMOTE_CACHE_ROOT)
            elif self.cloud_session is not None:
                download_sftp_project_index(self.cloud_session, self.remote_base_dir, REMOTE_CACHE_ROOT)
            return True
        except Exception as exc:
            self.remote_connected = False
            self.update_storage_status()
            messagebox.showerror(APP_TITLE, f"Не удалось обновить проекты из облака:\n{exc}")
            return False

    def refresh_remote_project(self, project_name: str) -> Optional[Path]:
        if not self.ensure_remote_ready():
            return None
        local_project = REMOTE_CACHE_ROOT / project_name
        try:
            if self.cloud_provider.get() == "yandex_disk":
                yandex_download_project_atomic(self.yandex_config.access_token, project_name, local_project)
            elif self.cloud_session is not None:
                remote_project = sftp_join(self.remote_base_dir, project_name)
                download_sftp_project_atomic(self.cloud_session, remote_project, local_project)
            return local_project
        except Exception as exc:
            self.remote_connected = False
            self.update_storage_status()
            messagebox.showerror(APP_TITLE, f"Не удалось обновить проект из облака:\n{exc}")
            return None

    def list_remote_project_files(self, project_name: str, mode: str) -> list[RemoteFile]:
        suffixes = ["_пресеты.xlsx", "_пресеты.pdf"] if mode == "presets" else ["_партитура.xlsx", "_партитура.pdf", "_new.xml"]
        if self.cloud_provider.get() == "yandex_disk":
            if not self.yandex_config.access_token:
                raise RuntimeError("Сначала подключите Яндекс.Диск.")
            remote_project = yandex_project_path(project_name)
            files = []
            for item in yandex_list_dir(self.yandex_config.access_token, remote_project):
                if item.get("type") != "file":
                    continue
                name = str(item.get("name") or "")
                if any(name.endswith(suffix) for suffix in suffixes):
                    files.append(RemoteFile(item.get("path") or f"{remote_project}/{name}", Path(name), int(item.get("size") or 0)))
            return sorted(files, key=lambda file: str(file.relative_path).lower())
        if self.cloud_session is None:
            raise RuntimeError("SFTP-сессия не открыта.")
        remote_project = sftp_join(self.remote_base_dir, project_name)
        files = []
        for item in self.cloud_session.listdir_attr(remote_project):
            if stat.S_ISDIR(item.st_mode):
                continue
            if any(item.filename.endswith(suffix) for suffix in suffixes):
                files.append(RemoteFile(sftp_join(remote_project, item.filename), Path(item.filename), int(item.st_size or 0)))
        return sorted(files, key=lambda file: str(file.relative_path).lower())

    def download_remote_file_to_cache(self, project_name: str, remote_file: RemoteFile) -> Path:
        target = REMOTE_CACHE_ROOT / project_name / remote_file.relative_path
        target.parent.mkdir(parents=True, exist_ok=True)
        if self.cloud_provider.get() == "yandex_disk":
            if not self.yandex_config.access_token:
                raise RuntimeError("Сначала подключите Яндекс.Диск.")
            target.write_bytes(yandex_get_bytes(yandex_download_url(self.yandex_config.access_token, remote_file.remote_path)))
            return target
        if self.cloud_session is None:
            raise RuntimeError("SFTP-сессия не открыта.")
        self.cloud_session.get(remote_file.remote_path, str(target))
        return target

    def upload_project_to_remote(self, project_dir: Path, ask_replace: bool = True) -> bool:
        if not self.ensure_remote_ready():
            return False
        require_project_xml(project_dir)
        if self.cloud_provider.get() == "yandex_disk":
            exists = (REMOTE_CACHE_ROOT / project_dir.name).exists()
            if exists and ask_replace and not messagebox.askyesno(APP_TITLE, "Такой проект уже есть в облаке. Заменить?"):
                return False
            yandex_upload_project(self.yandex_config.access_token, project_dir)
            return True
        if self.cloud_session is None:
            return False
        remote_project = sftp_join(self.remote_base_dir, project_dir.name)
        exists = sftp_exists(self.cloud_session, remote_project)
        if exists:
            if ask_replace and not messagebox.askyesno(APP_TITLE, "Такой проект уже есть в облаке. Заменить?"):
                return False
        upload_sftp_project_atomic(self.cloud_session, project_dir, remote_project)
        mirror_project_to_remote_cache(project_dir)
        return True

    def upload_project_to_remote_silent(
        self,
        project_dir: Path,
        progress: Optional[Callable[[int, int, str], None]] = None,
    ) -> None:
        require_project_xml(project_dir)
        if self.cloud_provider.get() == "yandex_disk":
            if not self.yandex_config.access_token:
                raise RuntimeError("Сначала подключите Яндекс.Диск.")
            yandex_upload_project(self.yandex_config.access_token, project_dir, progress=progress)
            self.remote_connected = True
            return
        config = self.sftp_config
        if not config.host or not config.username:
            raise RuntimeError("Сначала настройте подключение к облаку.")
        ssh = paramiko.SSHClient()
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        ssh.connect(
            hostname=config.host,
            port=config.port,
            username=config.username,
            password=config.password,
            timeout=12,
            banner_timeout=12,
            auth_timeout=12,
        )
        session = ssh.open_sftp()
        base = config.remote_dir.strip("/") or REMOTE_PROJECT_ROOT_NAME
        ensure_sftp_dir(session, base)
        remote_project = sftp_join(base, project_dir.name)
        upload_sftp_project_atomic(session, project_dir, remote_project, progress=progress)
        self.cloud_session = session
        self.ssh_client = ssh
        self.remote_base_dir = base
        self.remote_connected = True
        mirror_project_to_remote_cache(project_dir)

    def download_project_from_remote(self, project_name: str, local_root: Path, ask_replace: bool = True) -> Optional[Path]:
        if not self.ensure_remote_ready():
            return None
        local_project = local_root / project_name
        if local_project.exists():
            if ask_replace and not messagebox.askyesno(APP_TITLE, "Такой проект уже есть локально. Заменить?"):
                return None
        if self.cloud_provider.get() == "yandex_disk":
            yandex_download_project_atomic(self.yandex_config.access_token, project_name, local_project)
            return local_project
        if self.cloud_session is None:
            return None
        remote_project = sftp_join(self.remote_base_dir, project_name)
        if not sftp_exists(self.cloud_session, remote_project):
            messagebox.showerror(APP_TITLE, "В облаке проект не найден.")
            return None
        download_sftp_project_atomic(self.cloud_session, remote_project, local_project)
        return local_project

    def download_project_from_remote_silent(
        self,
        project_name: str,
        local_root: Path,
        progress: Optional[Callable[[int, int, str], None]] = None,
    ) -> Path:
        if self.cloud_provider.get() == "yandex_disk":
            if not self.yandex_config.access_token:
                raise RuntimeError("Сначала подключите Яндекс.Диск.")
            local_project = local_root / project_name
            yandex_download_project_atomic(self.yandex_config.access_token, project_name, local_project, progress=progress)
            self.remote_connected = True
            return local_project
        config = self.sftp_config
        if not config.host or not config.username:
            raise RuntimeError("Сначала настройте подключение к облаку.")
        ssh = paramiko.SSHClient()
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        ssh.connect(
            hostname=config.host,
            port=config.port,
            username=config.username,
            password=config.password,
            timeout=12,
            banner_timeout=12,
            auth_timeout=12,
        )
        session = ssh.open_sftp()
        base = config.remote_dir.strip("/") or REMOTE_PROJECT_ROOT_NAME
        ensure_sftp_dir(session, base)
        remote_project = sftp_join(base, project_name)
        if not sftp_exists(session, remote_project):
            raise RuntimeError("В облаке проект не найден.")
        local_project = local_root / project_name
        download_sftp_project_atomic(session, remote_project, local_project, progress=progress)
        self.cloud_session = session
        self.ssh_client = ssh
        self.remote_base_dir = base
        self.remote_connected = True
        return local_project

    def sync_current_project_to_remote(self) -> None:
        project = self.selected_project()
        if not project:
            return
        project_dir = project.directory
        self.show_transfer("Загружаю проект в облако...")

        def progress(done: int, total: int, name: str) -> None:
            text = f"Загружаю {done}/{total}: {name}"
            self.after(0, lambda: (self.storage_status.set(text), self.transfer_status.set(text)))

        def worker() -> None:
            try:
                self.upload_project_to_remote_silent(project_dir, progress=progress)
                ok = True
                self.after(0, lambda: self.on_remote_upload_done(ok, ""))
            except Exception as exc:
                self.after(0, lambda: self.on_remote_upload_done(False, str(exc)))

        threading.Thread(target=worker, daemon=True).start()

    def on_remote_upload_done(self, ok: bool, error: str) -> None:
        self.hide_transfer(return_to_previous=True)
        self.update_storage_status()
        if ok:
            if self.mode == "project_list" and self.storage_mode.get() == "remote":
                self.show_project_list(self.project_list_mode)
            messagebox.showinfo(APP_TITLE, "Проект сохранен в облако.")
        else:
            messagebox.showerror(APP_TITLE, f"Не удалось сохранить в облако:\n{error or 'ошибка подключения'}")

    def sync_current_project_to_local(self) -> None:
        project = self.selected_project()
        if not project:
            return
        self.copy_project_to_local_with_prompt(project)

    def copy_project_to_local_with_prompt(self, project: Project) -> None:
        try:
            if self.storage_mode.get() == "remote":
                target = PROJECT_ROOT / project.directory.name
                if target.exists() and not messagebox.askyesno(APP_TITLE, "Такой проект уже есть локально. Заменить?"):
                    return
                self.download_remote_project_in_background(project.directory.name)
                return
            else:
                target = PROJECT_ROOT / project.directory.name
                replace = True
                if target.exists():
                    replace = messagebox.askyesno(APP_TITLE, "Такой проект уже есть локально. Заменить?")
                copy_project_dir(project.directory, PROJECT_ROOT, replace=replace)
            messagebox.showinfo(APP_TITLE, "Проект сохранен локально.")
        except FileExistsError:
            messagebox.showwarning(APP_TITLE, "Проект уже есть локально.")
        except Exception as exc:
            messagebox.showerror(APP_TITLE, f"Не удалось сохранить локально:\n{exc}")

    def download_remote_project_in_background(self, project_name: str) -> None:
        self.show_transfer("Скачиваю проект из облака...")

        def progress(done: int, total: int, name: str) -> None:
            if total:
                text = f"Скачиваю {done}/{total}: {name}"
            else:
                text = f"Скачиваю: {name}"
            self.after(0, lambda: (self.storage_status.set(text), self.transfer_status.set(text)))

        def worker() -> None:
            try:
                target = self.download_project_from_remote_silent(project_name, PROJECT_ROOT, progress=progress)
                self.after(0, lambda: self.on_remote_download_done(True, target, ""))
            except Exception as exc:
                self.after(0, lambda: self.on_remote_download_done(False, None, str(exc)))

        threading.Thread(target=worker, daemon=True).start()

    def on_remote_download_done(self, ok: bool, target: Optional[Path], error: str) -> None:
        self.hide_transfer()
        self.update_storage_status()
        if ok and target is not None:
            set_active_project_root(PROJECT_ROOT)
            self.storage_mode.set("local")
            self.project = Project(project_title_from_dir(target), target, require_project_xml(target))
            self.projects = list_projects()
            messagebox.showinfo(APP_TITLE, "Проект сохранен локально.")
            self.show_project_mode()
        else:
            messagebox.showerror(APP_TITLE, f"Не удалось скачать проект:\n{error or 'ошибка подключения'}")

    def show_start(self) -> None:
        self.update_storage_status()
        self.show_frame("start")

    def show_project_source(self) -> None:
        self.update_storage_status()
        self.show_frame("project_source")

    def show_local_projects(self) -> None:
        self.storage_mode.set("local")
        set_active_project_root(PROJECT_ROOT)
        self.update_storage_status()
        self.show_project_list("projects")

    def show_remote_projects(self) -> None:
        self.storage_mode.set("remote")
        self.update_storage_status()
        self.refresh_remote_cache_in_background("projects")

    def show_preset_setup(self) -> None:
        self.show_frame("preset_setup")

    def show_partitura(self) -> None:
        self.show_frame("partitura")

    def back_from_project_list(self) -> None:
        self.show_project_source()

    def back_from_files(self) -> None:
        self.show_project_mode()

    def show_project_list(self, mode: str) -> None:
        if self.storage_mode.get() == "remote":
            set_active_project_root(REMOTE_CACHE_ROOT)
        else:
            set_active_project_root(PROJECT_ROOT)
        self.update_storage_status()
        self.project_list_mode = mode
        self.project_list_title.set("Проекты: облако" if self.storage_mode.get() == "remote" else "Проекты: устройство")
        if hasattr(self, "create_project_button"):
            if self.storage_mode.get() == "remote":
                self.create_project_button.grid_remove()
            else:
                self.create_project_button.grid()
                self.create_project_button.configure(state="normal")
        if self.storage_mode.get() == "remote":
            self.projects = getattr(self, "remote_projects", [])
        else:
            self.projects = list_projects()
        self.project_listbox.delete(0, "end")
        for project in self.projects:
            self.project_listbox.insert("end", project.title)
        self.show_frame("project_list")

    def refresh_remote_cache_in_background(self, mode: str) -> None:
        self.show_transfer("Получаю список проектов из облака...")

        def worker() -> None:
            try:
                if not self.ensure_remote_ready_silent():
                    raise RuntimeError("Не удалось подключиться к облаку.")
                if self.cloud_provider.get() == "yandex_disk":
                    names = yandex_project_names(self.yandex_config.access_token)
                elif self.cloud_session is None:
                    raise RuntimeError("SFTP-сессия не открыта.")
                else:
                    names = sftp_project_names(self.cloud_session, self.remote_base_dir)
                self.after(0, lambda: self.on_remote_cache_done(True, mode, "", names))
            except Exception as exc:
                self.after(0, lambda: self.on_remote_cache_done(False, mode, str(exc), []))

        threading.Thread(target=worker, daemon=True).start()

    def on_remote_cache_done(self, ok: bool, mode: str, error: str, names: list[str]) -> None:
        self.hide_transfer()
        self.update_storage_status()
        if ok:
            self.remote_projects = [
                Project(project_title_from_dir(REMOTE_CACHE_ROOT / name), REMOTE_CACHE_ROOT / name, REMOTE_CACHE_ROOT / name / f"{project_title_from_dir(REMOTE_CACHE_ROOT / name)}.xml")
                for name in names
            ]
            self.show_project_list(mode)
        else:
            messagebox.showerror(APP_TITLE, f"Не удалось обновить облако:\n{error}")
            self.show_project_source()

    def show_project_mode(self) -> None:
        if not self.project:
            self.show_project_list("projects")
            return
        self.project_mode_title.set(self.project.title)
        self.open_presets_button.configure(state="normal")
        self.open_partitura_button.configure(state="normal")
        self.show_frame("project_mode")

    def show_kind_menu(self, mode: str) -> None:
        if not self.project:
            return
        self.kind_menu.delete(0, "end")
        if self.storage_mode.get() != "remote":
            self.kind_menu.add_command(label="Открыть", command=lambda: self.open_project_builder(mode))
            self.kind_menu.add_command(label="Файлы", command=lambda: self.show_files(mode))
        if self.storage_mode.get() == "remote":
            self.kind_menu.add_command(label="Загрузить на устройство", command=lambda: self.copy_project_to_local_with_prompt(self.project))
        else:
            self.kind_menu.add_command(label="Загрузить в облако", command=self.sync_open_project_to_remote)
        self.kind_menu.tk_popup(self.winfo_pointerx(), self.winfo_pointery())

    def sync_open_project_to_remote(self) -> None:
        if not self.project:
            return
        project_dir = self.project.directory
        self.show_transfer("Загружаю проект в облако...")

        def progress(done: int, total: int, name: str) -> None:
            text = f"Загружаю {done}/{total}: {name}"
            self.after(0, lambda: (self.storage_status.set(text), self.transfer_status.set(text)))

        def worker() -> None:
            try:
                self.upload_project_to_remote_silent(project_dir, progress=progress)
                self.after(0, lambda: self.on_remote_upload_done(True, ""))
            except Exception as exc:
                self.after(0, lambda: self.on_remote_upload_done(False, str(exc)))

        threading.Thread(target=worker, daemon=True).start()

    def selected_project(self) -> Optional[Project]:
        selection = self.project_listbox.curselection()
        if not selection:
            messagebox.showwarning(APP_TITLE, "Выберите проект.")
            return None
        return self.projects[selection[0]]

    def show_project_menu(self, event) -> None:
        index = self.project_listbox.nearest(event.y)
        self.project_listbox.selection_clear(0, "end")
        self.project_listbox.selection_set(index)
        self.project_menu.delete(0, "end")
        self.project_menu.add_command(label="Открыть", command=self.open_selected_project)
        self.project_menu.add_command(label="Переименовать", command=self.rename_selected_project)
        if self.storage_mode.get() == "remote":
            self.project_menu.add_command(label="Загрузить на устройство", command=self.sync_current_project_to_local)
        else:
            self.project_menu.add_command(label="Загрузить в облако", command=self.sync_current_project_to_remote)
        self.project_menu.add_command(label="Удалить", command=self.delete_selected_project)
        self.project_menu.tk_popup(event.x_root, event.y_root)

    def open_selected_project(self) -> None:
        project = self.selected_project()
        if not project:
            return
        self.project = project
        self.show_project_mode()

    def open_project_builder(self, mode: str) -> None:
        if not self.project:
            return
        if self.storage_mode.get() == "remote":
            self.show_files(mode)
            return
        if mode == "partitura":
            self.partitura_project_var.set(f"Активный проект: {self.project.title}")
            self.load_partitura_state_for_project()
            self.show_partitura()
        else:
            self.open_preset_project(self.project)

    def open_selected_project_files(self) -> None:
        project = self.selected_project()
        if not project:
            return
        self.project = project
        self.show_files(self.project_list_mode)

    def rename_selected_project(self) -> None:
        project = self.selected_project()
        if not project:
            return
        new_title = simpledialog.askstring(APP_TITLE, "Новое название проекта:", initialvalue=project.title, parent=self)
        if new_title is None:
            return
        new_title = new_title.strip()
        if not new_title:
            return
        new_dir = project_dir_for_title(new_title)
        if new_dir.exists() and new_dir != project.directory:
            messagebox.showerror(APP_TITLE, "Проект с таким названием уже есть.")
            return
        if self.storage_mode.get() == "remote":
            try:
                old_name = project.directory.name
                new_name = new_dir.name
                if self.cloud_provider.get() == "yandex_disk":
                    yandex_rename_project(self.yandex_config.access_token, old_name, new_name)
                else:
                    if not self.ensure_remote_ready_silent() or self.cloud_session is None:
                        raise RuntimeError("Не удалось подключиться к облаку.")
                    self.cloud_session.rename(sftp_join(self.remote_base_dir, old_name), sftp_join(self.remote_base_dir, new_name))
                self.refresh_remote_cache_in_background(self.project_list_mode)
            except Exception as exc:
                messagebox.showerror(APP_TITLE, f"Не удалось переименовать в облаке:\n{exc}")
            return
        if project.directory != new_dir:
            project.directory.rename(new_dir)
        xml = project_xml_path(new_dir)
        if xml:
            target_xml = new_dir / f"{safe_filename(new_title)}.xml"
            if xml != target_xml:
                xml.rename(target_xml)
        if self.storage_mode.get() == "remote" and self.cloud_session is not None:
            remove_sftp_path(self.cloud_session, sftp_join(self.remote_base_dir, project.directory.name))
            self.upload_project_to_remote(new_dir, ask_replace=False)
        elif self.storage_mode.get() == "remote" and self.cloud_provider.get() == "yandex_disk":
            try:
                yandex_rename_project(self.yandex_config.access_token, project.directory.name, new_dir.name)
                mirror_project_to_remote_cache(new_dir)
            except Exception as exc:
                messagebox.showerror(APP_TITLE, f"Не удалось переименовать в облаке:\n{exc}")
        self.show_project_list(self.project_list_mode)

    def delete_selected_project(self) -> None:
        project = self.selected_project()
        if not project:
            return
        if messagebox.askyesno(APP_TITLE, f"Удалить проект «{project.title}» целиком?"):
            if self.storage_mode.get() == "remote":
                try:
                    if self.cloud_provider.get() == "yandex_disk":
                        yandex_delete_project(self.yandex_config.access_token, project.directory.name)
                    else:
                        if not self.ensure_remote_ready_silent() or self.cloud_session is None:
                            raise RuntimeError("Не удалось подключиться к облаку.")
                        remove_sftp_path(self.cloud_session, sftp_join(self.remote_base_dir, project.directory.name))
                except Exception as exc:
                    messagebox.showerror(APP_TITLE, f"Не удалось удалить в облаке:\n{exc}")
                self.refresh_remote_cache_in_background(self.project_list_mode)
                return
            shutil.rmtree(project.directory)
            self.show_project_list(self.project_list_mode)

    def create_preset_project_from_xml(self) -> None:
        project = self.create_project_from_xml_dialog()
        if project:
            self.open_preset_project(project)

    def create_partitura_project_from_xml(self) -> None:
        project = self.create_project_from_xml_dialog()
        if project:
            self.project = project
            self.partitura_project_var.set(f"Активный проект: {project.title}")
            self.load_partitura_state_for_project()
            self.show_partitura()

    def create_project_from_xml_dialog(self) -> Optional[Project]:
        filename = filedialog.askopenfilename(title="Выберите XML GrandMA2", filetypes=[("GrandMA2 XML", "*.xml"), ("Все файлы", "*.*")])
        if not filename:
            return None
        source = Path(filename)
        title = simpledialog.askstring(APP_TITLE, "Название спектакля:", initialvalue=source.stem, parent=self)
        if title is None:
            return None
        title = title.strip() or source.stem
        try:
            project_dir = project_dir_for_title(title)
            if project_dir.exists():
                if not messagebox.askyesno(APP_TITLE, "Такой проект уже есть. Заменить?"):
                    return None
                shutil.rmtree(project_dir)
            project = copy_xml_to_project(source, title)
            return project
        except Exception as exc:
            messagebox.showerror(APP_TITLE, f"Не удалось создать проект:\n{exc}")
            return None

    def open_preset_project(self, project: Project) -> None:
        try:
            items = parse_grandma2_presets(project.xml_path)
            if not items:
                messagebox.showwarning(APP_TITLE, "В XML не нашлось пресетов с приборами.")
                return
            rows, warning = load_passport_rows(items, project.directory, project.title)
        except Exception as exc:
            messagebox.showerror(APP_TITLE, f"Не удалось открыть проект:\n{exc}")
            return
        self.project = project
        self.items = items
        self.rows = rows
        self.index = 0
        self.reviewing_photo = False
        self.workspace_summary.set(f"{project.title}: пресетов {len({(i.preset_no, i.preset_name) for i in items})}, строк {len(rows)}")
        self.refresh_table()
        self.show_frame("workspace")
        self.show_current_item()
        self.autosave_passport()
        self.connect_camera(silent=True)
        if warning:
            messagebox.showwarning(APP_TITLE, warning)

    def show_files(self, mode: str) -> None:
        if not self.project:
            self.show_project_list(mode)
            return
        self.files_mode = mode
        self.files_title.set(f"Файлы: {self.project.title}")
        self.files_listbox.delete(0, "end")
        self.current_files = []
        self.current_remote_files = []
        suffixes = ["_пресеты.xlsx", "_пресеты.pdf"] if mode == "presets" else ["_партитура.xlsx", "_партитура.pdf", "_new.xml"]
        if self.storage_mode.get() == "remote":
            try:
                if not self.ensure_remote_ready_silent():
                    raise RuntimeError("Не удалось подключиться к облаку.")
                self.current_remote_files = self.list_remote_project_files(self.project.directory.name, mode)
            except Exception as exc:
                messagebox.showerror(APP_TITLE, f"Не удалось получить файлы из облака:\n{exc}")
                return
            for remote_file in self.current_remote_files:
                label = str(remote_file.relative_path)
                if remote_file.size:
                    label += f"  ({readable_size(remote_file.size)})"
                self.files_listbox.insert("end", label)
            self.show_frame("files")
            return
        for path in sorted(self.project.directory.iterdir(), key=lambda p: p.name.lower()):
            if path.is_file() and any(path.name.endswith(suffix) for suffix in suffixes):
                self.current_files.append(path)
                self.files_listbox.insert("end", path.name)
        self.show_frame("files")

    def selected_file(self) -> Optional[Path]:
        selection = self.files_listbox.curselection()
        if not selection:
            messagebox.showwarning(APP_TITLE, "Выберите файл.")
            return None
        return self.current_files[selection[0]]

    def show_file_menu(self, event) -> None:
        index = self.files_listbox.nearest(event.y)
        self.files_listbox.selection_clear(0, "end")
        self.files_listbox.selection_set(index)
        self.file_menu.tk_popup(event.x_root, event.y_root)

    def open_selected_file(self) -> None:
        selection = self.files_listbox.curselection()
        if not selection:
            messagebox.showwarning(APP_TITLE, "Выберите файл.")
            return
        if self.storage_mode.get() == "remote" and self.project:
            try:
                remote_file = self.current_remote_files[selection[0]]
                path = self.download_remote_file_to_cache(self.project.directory.name, remote_file)
            except Exception as exc:
                messagebox.showerror(APP_TITLE, f"Не удалось скачать файл:\n{exc}")
                return
            open_path(path)
            return
        path = self.current_files[selection[0]]
        open_path(path)

    def reveal_selected_file(self) -> None:
        path = self.selected_file()
        if path:
            reveal_path(path)

    def delete_selected_file(self) -> None:
        path = self.selected_file()
        if path and messagebox.askyesno(APP_TITLE, f"Удалить файл {path.name}?"):
            path.unlink()
            if self.storage_mode.get() == "remote" and self.project:
                self.upload_project_to_remote(self.project.directory, ask_replace=False)
            self.show_files(self.files_mode)

    def refresh_table(self) -> None:
        self.table.delete(*self.table.get_children())
        for index, row in enumerate(self.rows):
            self.table.insert("", "end", iid=str(index), values=self.table_values(row))

    def table_values(self, row: PassportRow) -> tuple[str, str, str, str]:
        mark = "✓" if row.photo_path and row.photo_path.exists() else ""
        desc = row.description.replace("\n", " ")
        return mark, row.preset_label, row.fixture_id, desc

    def mark_table_row(self, row_index: int) -> None:
        if 0 <= row_index < len(self.rows):
            self.table.item(str(row_index), values=self.table_values(self.rows[row_index]))

    def show_current_item(self) -> None:
        if not self.rows:
            return
        self.index = max(0, min(self.index, len(self.rows) - 1))
        row = self.rows[self.index]
        self.current_var.set(f"{self.index + 1}/{len(self.rows)}  Пресет: {row.preset_label}   Прибор: {row.fixture_id}")
        self.loading_description = True
        self.description_text.set_real_text(row.description)
        self.loading_description = False
        self.syncing_selection = True
        self.table.selection_set(str(self.index))
        self.table.focus(str(self.index))
        self.table.see(str(self.index))
        self.after_idle(lambda: setattr(self, "syncing_selection", False))
        self.update_progress()
        self.update_photo_area()

    def update_photo_area(self) -> None:
        self.delete_photo_button.place_forget()
        self.hide_review_buttons()
        row = self.rows[self.index]
        if row.photo_path and row.photo_path.exists():
            self.reviewing_photo = True
            self.show_photo_preview(row.photo_path)
            self.delete_photo_button.place(x=12, y=12)
            self.photo_button.configure(text="Переснять", state="normal")
            self.add_row_button.grid()
        else:
            self.reviewing_photo = False
            self.video_label.configure(image="", text="Фото еще нет\nили нажми кнопку «Фото» внизу")
            self.preview_image = None
            self.photo_button.configure(text="Фото", state="normal")
            self.add_row_button.grid_remove()

    def on_table_select(self, _event=None) -> None:
        if self.syncing_selection:
            return
        selection = self.table.selection()
        if not selection:
            return
        self.save_current_description()
        self.index = int(selection[0])
        self.clear_capture()
        self.show_current_item()

    def show_table_menu(self, event) -> None:
        row_id = self.table.identify_row(event.y)
        if not row_id:
            return
        self.table.selection_set(row_id)
        self.on_table_select()
        self.table_menu.tk_popup(event.x_root, event.y_root)

    def on_description_changed(self, _event=None) -> None:
        if self.loading_description or not self.rows:
            return
        self.rows[self.index].description = self.description_text.real_text()
        self.mark_table_row(self.index)
        self.autosave_passport()

    def save_current_description(self) -> None:
        if self.rows:
            self.rows[self.index].description = self.description_text.real_text()
            self.mark_table_row(self.index)

    def connect_camera(self, silent: bool = False) -> None:
        if self.camera_opening:
            return
        self.stop_camera()
        source = source_from_text(self.camera_source_var.get())
        self.camera_opening = True
        self.camera_queue = multiprocessing.get_context("spawn").Queue(maxsize=3)
        self.camera_stop_event = multiprocessing.get_context("spawn").Event()
        self.camera_process = multiprocessing.get_context("spawn").Process(
            target=camera_worker,
            args=(source, self.camera_queue, self.camera_stop_event),
            daemon=True,
        )
        self.camera_process.start()
        if not silent:
            self.workspace_summary.set("Подключаю камеру...")
        self.after(50, lambda: self.poll_camera(silent))

    def poll_camera(self, silent: bool = False) -> None:
        if self.camera_queue is None:
            return
        got_frame = False
        try:
            while True:
                message = self.camera_queue.get_nowait()
                if message[0] == "connected":
                    self.camera_opening = False
                    self.camera_running = True
                    if not silent and self.project:
                        self.workspace_summary.set(f"{self.project.title}: камера подключена ({message[1]})")
                elif message[0] == "frame":
                    self.current_frame_bytes = message[1]
                    self.camera_last_frame_time = message[2]
                    got_frame = True
                elif message[0] == "error":
                    self.handle_camera_error(message[1], silent)
                    return
        except queue.Empty:
            pass
        if got_frame and not self.reviewing_photo and self.current_frame_bytes is not None and self.mode == "workspace":
            self.show_frame_bytes(self.current_frame_bytes)
        if self.camera_process is not None and self.camera_process.exitcode is not None and self.camera_process.exitcode != 0:
            self.handle_camera_error(f"Драйвер камеры упал, код {self.camera_process.exitcode}", silent)
            return
        if self.camera_running or self.camera_opening:
            self.after(30, lambda: self.poll_camera(silent))

    def handle_camera_error(self, detail: str, silent: bool = False) -> None:
        self.stop_camera()
        if not silent:
            messagebox.showerror(
                APP_TITLE,
                f"{detail}.\n\nПопробуйте другой источник камеры: 0, 1, 2, 3.",
            )

    def stop_camera(self) -> None:
        self.camera_opening = False
        self.camera_running = False
        if self.camera_stop_event is not None:
            self.camera_stop_event.set()
        if self.camera_process is not None and self.camera_process.is_alive():
            self.camera_process.join(timeout=0.4)
            if self.camera_process.is_alive():
                self.camera_process.terminate()
        self.camera_process = None
        self.camera_queue = None
        self.camera_stop_event = None

    def apply_zoom(self, image: Image.Image) -> Image.Image:
        try:
            zoom = max(1.0, float(self.zoom_var.get()))
        except tk.TclError:
            zoom = 1.0
        if zoom <= 1.01:
            return image
        width, height = image.size
        crop_w = max(1, int(width / zoom))
        crop_h = max(1, int(height / zoom))
        left = (width - crop_w) // 2
        top = (height - crop_h) // 2
        return image.crop((left, top, left + crop_w, top + crop_h))

    def fit_to_video_label(self, image: Image.Image) -> ImageTk.PhotoImage:
        label_width = max(self.video_label.winfo_width(), 640)
        label_height = max(self.video_label.winfo_height(), 360)
        image.thumbnail((label_width, label_height))
        return ImageTk.PhotoImage(image)

    def show_frame_bytes(self, frame_bytes: bytes) -> None:
        image = Image.open(io.BytesIO(frame_bytes)).convert("RGB")
        image = self.apply_zoom(image)
        self.preview_image = self.fit_to_video_label(image)
        self.video_label.configure(image=self.preview_image, text="")

    def show_photo_preview(self, path: Path) -> None:
        image = Image.open(path).convert("RGB")
        self.preview_image = self.fit_to_video_label(image)
        self.video_label.configure(image=self.preview_image, text="")

    def take_photo(self) -> None:
        if self.rows[self.index].photo_path and self.rows[self.index].photo_path.exists():
            self.delete_current_photo(autosave=False)
        if self.current_frame_bytes is None:
            messagebox.showwarning(APP_TITLE, "Нет кадра с камеры.")
            return
        image = Image.open(io.BytesIO(self.current_frame_bytes)).convert("RGB")
        image = self.apply_zoom(image)
        tmp_dir = Path(tempfile.gettempdir()) / "passport_creator"
        tmp_dir.mkdir(parents=True, exist_ok=True)
        self.captured_temp = tmp_dir / f"capture_{int(time.time() * 1000)}.jpg"
        image.save(self.captured_temp, quality=94)
        self.reviewing_photo = True
        self.show_photo_preview(self.captured_temp)
        self.show_review_buttons()

    def show_review_buttons(self) -> None:
        self.photo_button.grid_remove()
        self.upload_photo_button.grid_remove()
        self.add_row_button.grid_remove()
        self.use_button.grid()
        self.retake_button.grid()

    def hide_review_buttons(self) -> None:
        self.use_button.grid_remove()
        self.retake_button.grid_remove()
        self.photo_button.grid()
        self.upload_photo_button.grid()

    def refresh_zoom_preview(self) -> None:
        if self.reviewing_photo:
            return
        if self.current_frame_bytes is not None and self.mode == "workspace":
            self.show_frame_bytes(self.current_frame_bytes)

    def use_photo(self) -> None:
        if self.captured_temp is None or self.project is None:
            return
        self.save_current_description()
        target = self.unique_photo_path(self.rows[self.index].file_stem)
        if self.rows[self.index].photo_path:
            self.delete_photo_file(self.rows[self.index].photo_path)
        shutil.move(str(self.captured_temp), target)
        self.rows[self.index].photo_path = target
        self.captured_temp = None
        self.mark_table_row(self.index)
        self.autosave_passport()
        self.index = self.next_index_after_photo()
        self.show_current_item()

    def retake_photo(self) -> None:
        self.clear_capture()
        self.reviewing_photo = False
        if self.current_frame_bytes:
            self.show_frame_bytes(self.current_frame_bytes)
        self.hide_review_buttons()

    def next_index_after_photo(self) -> int:
        if not self.rows:
            return 0
        for candidate in range(self.index + 1, len(self.rows)):
            row = self.rows[candidate]
            if not row.photo_path or not row.photo_path.exists():
                return candidate
        for candidate in range(0, self.index):
            row = self.rows[candidate]
            if not row.photo_path or not row.photo_path.exists():
                return candidate
        return min(self.index + 1, len(self.rows) - 1)

    def clear_capture(self) -> None:
        if self.captured_temp and self.captured_temp.exists():
            self.captured_temp.unlink()
        self.captured_temp = None

    def unique_photo_path(self, stem: str) -> Path:
        assert self.project is not None
        photo_dir = photos_dir(self.project.directory)
        path = photo_dir / f"{stem}.jpg"
        counter = 2
        while path.exists():
            path = photo_dir / f"{stem}_{counter}.jpg"
            counter += 1
        return path

    def upload_photo(self) -> None:
        filename = filedialog.askopenfilename(
            title="Выберите фото",
            filetypes=[("Фото", "*.jpg *.jpeg *.png *.webp *.bmp *.heic *.heif"), ("Все файлы", "*.*")],
        )
        if not filename:
            return
        source = Path(filename)
        if source.suffix.lower() not in PHOTO_EXTENSIONS:
            messagebox.showwarning(APP_TITLE, "Выберите файл изображения.")
            return
        target = self.unique_photo_path(self.rows[self.index].file_stem)
        with Image.open(source) as image:
            image.convert("RGB").save(target, quality=94)
        self.rows[self.index].photo_path = target
        self.autosave_passport()
        self.show_current_item()

    def add_extra_row(self) -> None:
        self.save_current_description()
        base = self.rows[self.index]
        new_row = PassportRow(base.preset_label, base.fixture_id, base.preset_no)
        self.rows.insert(self.index + 1, new_row)
        self.refresh_table()
        self.index += 1
        self.autosave_passport()
        self.show_current_item()

    def delete_current_photo(self, autosave: bool = True) -> None:
        if not self.rows:
            return
        self.delete_photo_file(self.rows[self.index].photo_path)
        self.rows[self.index].photo_path = None
        self.mark_table_row(self.index)
        if autosave:
            self.autosave_passport()
        self.show_current_item()

    def delete_current_row(self) -> None:
        if len(self.rows) <= 1:
            messagebox.showwarning(APP_TITLE, "Последнюю запись удалить нельзя.")
            return
        row = self.rows[self.index]
        if not messagebox.askyesno(APP_TITLE, f"Удалить запись {row.preset_label} | {row.fixture_id}?"):
            return
        self.delete_photo_file(row.photo_path)
        del self.rows[self.index]
        self.index = min(self.index, len(self.rows) - 1)
        self.refresh_table()
        self.autosave_passport()
        self.show_current_item()

    def delete_photo_file(self, path: Optional[Path]) -> None:
        if path and path.exists():
            path.unlink()

    def update_progress(self) -> None:
        ready = sum(1 for row in self.rows if row.photo_path and row.photo_path.exists())
        self.progress_var.set(f"{ready} / {len(self.rows)} с фото")

    def autosave_passport(self) -> None:
        if not self.project:
            return
        try:
            create_passport_xlsx(self.rows, preset_xlsx_path(self.project.directory, self.project.title), self.project.title)
            create_passport_pdf(self.rows, preset_pdf_path(self.project.directory, self.project.title), self.project.title)
        except Exception as exc:
            self.workspace_summary.set(f"Ошибка автосохранения: {exc}")

    def refresh_partitura_fields(self) -> None:
        self.field_listbox.delete(0, "end")
        for field in self.partitura_fields:
            mark = "✓" if field.enabled else " "
            self.field_listbox.insert("end", f"[{mark}] {field.title}")
            if not field.enabled:
                self.field_listbox.itemconfig("end", fg=MUTED)

    def selected_partitura_field_index(self) -> Optional[int]:
        selection = self.field_listbox.curselection()
        return selection[0] if selection else None

    def toggle_partitura_field(self) -> None:
        index = self.selected_partitura_field_index()
        if index is None:
            return
        self.partitura_fields[index].enabled = not self.partitura_fields[index].enabled
        self.refresh_partitura_fields()
        self.field_listbox.selection_set(index)

    def move_partitura_field(self, direction: int) -> None:
        index = self.selected_partitura_field_index()
        if index is None:
            return
        new_index = index + direction
        if not 0 <= new_index < len(self.partitura_fields):
            return
        self.partitura_fields[index], self.partitura_fields[new_index] = self.partitura_fields[new_index], self.partitura_fields[index]
        self.refresh_partitura_fields()
        self.field_listbox.selection_set(new_index)

    def create_partitura_files(self) -> None:
        if not self.project:
            messagebox.showwarning(APP_TITLE, "Сначала выберите или создайте проект.")
            return
        fields = [field for field in self.partitura_fields if field.enabled]
        if not fields:
            messagebox.showwarning(APP_TITLE, "Включите хотя бы одно поле.")
            return
        try:
            if not self.partitura_rows:
                self.load_partitura_state_for_project()
            rows = rows_for_original_partitura(self.partitura_rows, self.partitura_translated)
            self.partitura_rows = rows
            self.partitura_translated = False
            create_partitura_xlsx(rows, fields, partitura_xlsx_path(self.project.directory, self.project.title), self.project.title)
            create_partitura_pdf(rows, fields, partitura_pdf_path(self.project.directory, self.project.title), self.project.title)
            write_partitura_state(self.project.directory, self.partitura_rows, False)
        except Exception as exc:
            messagebox.showerror(APP_TITLE, f"Не удалось создать партитуру:\n{exc}")
            return
        self.show_files("partitura")

    def load_partitura_state_for_project(self) -> None:
        if not self.project:
            self.partitura_rows = []
            self.partitura_translated = False
            return
        self.partitura_rows, self.partitura_translated = load_partitura_rows(self.project)
        self.partitura_project_var.set(
            f"Активный проект: {self.project.title}"
        )

    def save_partitura_show_xml(self) -> None:
        if not self.project:
            messagebox.showwarning(APP_TITLE, "Сначала выберите проект.")
            return
        try:
            if not self.partitura_rows:
                self.load_partitura_state_for_project()
            write_partitura_state(self.project.directory, self.partitura_rows, self.partitura_translated)
            output = partitura_new_xml_path(self.project.directory, self.project.title)
            save_partitura_to_show_xml(self.project.xml_path, self.partitura_rows, output)
            messagebox.showinfo(APP_TITLE, f"Создан файл: {output.name}")
            self.show_files("partitura")
        except Exception as exc:
            messagebox.showerror(APP_TITLE, f"Не удалось сохранить show-файл:\n{exc}")

    def on_close(self) -> None:
        self.save_current_description()
        self.autosave_passport()
        self.stop_camera()
        self.close_remote()
        self.destroy()


def create_cell_photo(source: Path, target: Path, width: int, height: int) -> None:
    with Image.open(source) as image:
        image = image.convert("RGB")
        image.thumbnail((width, height))
        canvas = Image.new("RGB", (width, height), "white")
        left = (width - image.width) // 2
        top = (height - image.height) // 2
        canvas.paste(image, (left, top))
        canvas.save(target)


def create_passport_xlsx(rows: list[PassportRow], path: Path, title: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Паспорт"
    ws.append([title])
    ws.merge_cells("A1:D1")
    ws.append(["Пресет", "Прибор", "Фото", "Описание"])

    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws["A1"].font = Font(bold=True, size=16)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30
    for cell in ws[2]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        cell.fill = PatternFill("solid", fgColor="EEEEEE")

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 42
    ws.column_dimensions["D"].width = 34
    photo_width = 300
    photo_height = 180

    with tempfile.TemporaryDirectory() as temp_dir:
        prepared_photos: list[Path] = []
        for row_index, row in enumerate(rows, start=3):
            ws.cell(row=row_index, column=1, value=row.preset_label)
            ws.cell(row=row_index, column=2, value=row.fixture_id)
            ws.cell(row=row_index, column=4, value=row.description)
            ws.row_dimensions[row_index].height = 138
            for col in range(1, 5):
                cell = ws.cell(row=row_index, column=col)
                cell.border = border
                cell.alignment = Alignment(horizontal="center" if col != 4 else "left", vertical="center", wrap_text=True)
            if row.photo_path and row.photo_path.exists():
                prepared = Path(temp_dir) / f"photo_{row_index}.png"
                create_cell_photo(row.photo_path, prepared, photo_width, photo_height)
                prepared_photos.append(prepared)
                img = XlsxImage(str(prepared))
                img.width = photo_width
                img.height = photo_height
                ws.add_image(img, f"C{row_index}")

        merge_duplicate_passport_cells(ws, rows)
        ws.freeze_panes = "A3"
        wb.save(path)


def merge_duplicate_passport_cells(ws, rows: list[PassportRow]) -> None:
    start = 3
    while start < len(rows) + 3:
        key = rows[start - 3].group_key
        end = start
        while end + 1 < len(rows) + 3 and rows[end - 2].group_key == key:
            end += 1
        if end > start:
            ws.merge_cells(start_row=start, start_column=1, end_row=end, end_column=1)
            ws.merge_cells(start_row=start, start_column=2, end_row=end, end_column=2)
        start = end + 1


def create_partitura_xlsx(rows: list[PartituraRow], fields: list[PartituraField], path: Path, title: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Партитура"
    last_col = max(1, len(fields))
    ws.append([title])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    ws.append([field.title for field in fields])

    thin = Side(style="thin", color="555555")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws["A1"].font = Font(bold=True, size=16)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    for cell in ws[2]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    widths = {
        "number": 12,
        "name": 48,
        "trigger": 13,
        "trigger_time": 15,
        "fade": 10,
        "downfade": 12,
        "delay": 10,
        "info": 54,
        "command": 32,
    }
    for index, field in enumerate(fields, start=1):
        ws.column_dimensions[column_letter(index)].width = widths.get(field.field_id, 18)

    for row in rows:
        ws.append([row.value(field.field_id) for field in fields])

    for sheet_row in ws.iter_rows(min_row=3):
        for col_index, cell in enumerate(sheet_row, start=1):
            field = fields[col_index - 1]
            align = "left" if field.field_id in {"name", "info", "command"} else "center"
            cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
            cell.border = border

    ws.freeze_panes = "A3"
    wb.save(path)


def column_letter(index: int) -> str:
    result = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        result = chr(65 + remainder) + result
    return result


def create_passport_pdf(rows: list[PassportRow], path: Path, title: str) -> None:
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import landscape, A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.platypus import Image as PdfImage
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except ImportError:
        create_simple_pdf_placeholder(path, title, "Установите reportlab для PDF.")
        return

    font_name, bold_font_name = register_pdf_fonts()
    doc = SimpleDocTemplate(str(path), pagesize=landscape(A4), rightMargin=12 * mm, leftMargin=12 * mm, topMargin=10 * mm, bottomMargin=10 * mm)
    styles = getSampleStyleSheet()
    styles["Title"].fontName = bold_font_name
    normal = ParagraphStyle("normal", parent=styles["Normal"], fontName=font_name, fontSize=8, leading=10, alignment=1)
    left = ParagraphStyle("left", parent=normal, alignment=0)
    story = [Paragraph(title, styles["Title"]), Spacer(1, 5 * mm)]
    data = [["Пресет", "Прибор", "Фото", "Описание"]]
    temp_files: list[Path] = []
    for row in rows:
        photo = ""
        if row.photo_path and row.photo_path.exists():
            prepared = Path(tempfile.gettempdir()) / f"passport_pdf_{time.time_ns()}.jpg"
            create_cell_photo(row.photo_path, prepared, 240, 145)
            temp_files.append(prepared)
            photo = PdfImage(str(prepared), width=64 * mm, height=38 * mm)
        data.append([Paragraph(row.preset_label, normal), Paragraph(row.fixture_id, normal), photo, Paragraph(row.description, left)])
    table = Table(data, colWidths=[62 * mm, 22 * mm, 70 * mm, 90 * mm], repeatRows=1)
    style_commands = [
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("BACKGROUND", (0, 0), (-1, 0), colors.whitesmoke),
        ("FONTNAME", (0, 0), (-1, 0), bold_font_name),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]
    start = 1
    while start <= len(rows):
        key = rows[start - 1].group_key
        end = start
        while end < len(rows) and rows[end].group_key == key:
            end += 1
        if end > start:
            style_commands.append(("SPAN", (0, start), (0, end)))
            style_commands.append(("SPAN", (1, start), (1, end)))
        start = end + 1
    table.setStyle(TableStyle(style_commands))
    story.append(table)
    doc.build(story)
    for temp in temp_files:
        temp.unlink(missing_ok=True)


def create_partitura_pdf(rows: list[PartituraRow], fields: list[PartituraField], path: Path, title: str) -> None:
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except ImportError:
        create_simple_pdf_placeholder(path, title, "Установите reportlab для PDF.")
        return

    font_name, bold_font_name = register_pdf_fonts()
    doc = SimpleDocTemplate(str(path), pagesize=A4, rightMargin=10 * mm, leftMargin=10 * mm, topMargin=10 * mm, bottomMargin=10 * mm)
    styles = getSampleStyleSheet()
    styles["Title"].fontName = bold_font_name
    small = ParagraphStyle("small", parent=styles["Normal"], fontName=font_name, fontSize=6.4, leading=7.4)
    header = ParagraphStyle("header", parent=small, fontName=bold_font_name, alignment=1)
    story = [Paragraph(title, styles["Title"]), Spacer(1, 4 * mm)]
    data = [[Paragraph(field.title, header) for field in fields]]
    for row in rows:
        data.append([Paragraph(row.value(field.field_id).replace("\n", "<br/>"), small) for field in fields])
    weights = [partitura_pdf_weight(field.field_id) for field in fields]
    total = sum(weights) or 1
    width = 190 * mm
    col_widths = [width * weight / total for weight in weights]
    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.45, colors.black),
        ("BACKGROUND", (0, 0), (-1, 0), colors.white),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    story.append(table)
    doc.build(story)


def partitura_pdf_weight(field_id: str) -> float:
    return {
        "number": 0.75,
        "name": 2.4,
        "trigger": 0.9,
        "trigger_time": 1.05,
        "fade": 0.65,
        "downfade": 0.8,
        "delay": 0.65,
        "info": 2.6,
        "command": 1.4,
    }.get(field_id, 1.0)


def register_pdf_fonts() -> tuple[str, str]:
    try:
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
    except ImportError:
        return "Helvetica", "Helvetica-Bold"

    candidates = [
        (
            Path("/System/Library/Fonts/Supplemental/Arial.ttf"),
            Path("/System/Library/Fonts/Supplemental/Arial Bold.ttf"),
        ),
        (
            Path("/Library/Fonts/Arial.ttf"),
            Path("/Library/Fonts/Arial Bold.ttf"),
        ),
        (
            Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
            Path("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"),
        ),
    ]
    for regular, bold in candidates:
        if regular.exists() and bold.exists():
            try:
                if "PassportArial" not in pdfmetrics.getRegisteredFontNames():
                    pdfmetrics.registerFont(TTFont("PassportArial", str(regular)))
                if "PassportArialBold" not in pdfmetrics.getRegisteredFontNames():
                    pdfmetrics.registerFont(TTFont("PassportArialBold", str(bold)))
                return "PassportArial", "PassportArialBold"
            except Exception:
                continue
    return "Helvetica", "Helvetica-Bold"


def create_simple_pdf_placeholder(path: Path, title: str, message: str) -> None:
    image = Image.new("RGB", (1240, 1754), "white")
    draw = ImageDraw.Draw(image)
    font = ImageFont.load_default()
    draw.text((80, 80), title, fill="black", font=font)
    draw.text((80, 130), message, fill="black", font=font)
    image.save(path, "PDF")


def open_path(path: Path) -> None:
    if sys.platform == "darwin":
        subprocess.run(["open", str(path)], check=False)
    elif os.name == "nt":
        os.startfile(path)  # type: ignore[attr-defined]
    else:
        subprocess.run(["xdg-open", str(path)], check=False)


def readable_size(size: int) -> str:
    value = float(size)
    for unit in ("Б", "КБ", "МБ", "ГБ"):
        if value < 1024 or unit == "ГБ":
            return f"{value:.1f} {unit}" if unit != "Б" else f"{int(value)} {unit}"
        value /= 1024
    return f"{size} Б"


def reveal_path(path: Path) -> None:
    if sys.platform == "darwin":
        subprocess.run(["open", "-R", str(path)], check=False)
    elif os.name == "nt":
        subprocess.run(["explorer", f"/select,{path}"], check=False)
    else:
        open_path(path.parent)


def print_xml_summary(xml_path: Path) -> None:
    items = parse_grandma2_presets(xml_path)
    print(f"XML: {xml_path}")
    print(f"Пресетов: {len({(item.preset_no, item.preset_name) for item in items})}")
    print(f"Строк пресет-прибор: {len(items)}")
    for item in items[:10]:
        print(f"  {item.preset_label} | {item.preset_no} | fixture {item.fixture_id}")
    if len(items) > 10:
        print("  ...")


def print_camera_check() -> None:
    context = multiprocessing.get_context("spawn")
    for index in range(6):
        frame_queue = context.Queue(maxsize=3)
        stop_event = context.Event()
        process = context.Process(target=camera_worker, args=(index, frame_queue, stop_event), daemon=True)
        process.start()
        got_frame = False
        error = ""
        deadline = time.time() + 5
        while time.time() < deadline:
            try:
                message = frame_queue.get(timeout=0.2)
            except queue.Empty:
                continue
            if message[0] == "frame":
                got_frame = True
                break
            if message[0] == "error":
                error = message[1]
                break
        stop_event.set()
        process.join(timeout=0.5)
        if process.is_alive():
            process.terminate()
        print(f"{index}: {'OK' if got_frame else 'нет кадра ' + error}".strip())


if __name__ == "__main__":
    multiprocessing.freeze_support()
    if len(sys.argv) == 3 and sys.argv[1] == "--check":
        print_xml_summary(Path(sys.argv[2]))
    elif len(sys.argv) == 2 and sys.argv[1] == "--camera-check":
        print_camera_check()
    else:
        app = PassportApp()
        app.mainloop()
