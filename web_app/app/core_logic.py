from __future__ import annotations

import io
import multiprocessing
import os
import queue
import re
import shutil
import stat
import subprocess
import sys
import tempfile
import threading
import time
import copy
from dataclasses import dataclass
from pathlib import Path
from typing import Callable, Optional
from zipfile import BadZipFile
import json
import xml.etree.ElementTree as ET

# Core required libraries
from PIL import Image, ImageDraw, ImageFont
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XlsxImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter as column_letter

# Optional/Environment-specific libraries
try:
    import paramiko
except ImportError:
    paramiko = None

try:
    from PIL import ImageTk
except ImportError:
    ImageTk = None



APP_TITLE = "Passport creator"
PROJECT_ROOT = Path.home() / "Documents" / "MA2_passports"
APP_DATA_DIR = Path.home() / ".passport_creator"
REMOTE_CACHE_ROOT = APP_DATA_DIR / "remote_cache" / "MA2_passports"
CONFIG_PATH = APP_DATA_DIR / "cloud_connection.json"
ACTIVE_PROJECT_ROOT = PROJECT_ROOT
REMOTE_PROJECT_ROOT_NAME = "MA2_passports"
PHOTO_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp", ".bmp", ".heic", ".heif"}
XML_EXTENSIONS = {".xml"}

BLACK = "#000000"
PANEL = "#141414"
PANEL_2 = "#202020"
YELLOW = "#ffb800"
SILVER = "#cfcfcf"
MUTED = "#9c9c9c"
RED = "#ff4545"
GREEN = "#4ee06d"


@dataclass(frozen=True)
class PresetItem:
    preset_no: str
    preset_name: str
    fixture_id: str

    @property
    def preset_label(self) -> str:
        return self.preset_name.strip() or self.preset_no.strip()

    @property
    def file_stem(self) -> str:
        return safe_filename(f"{self.preset_no}_{self.fixture_id}")


@dataclass
class PassportRow:
    preset_label: str
    fixture_id: str
    preset_no: str
    photo_path: Optional[Path] = None
    description: str = ""

    @property
    def group_key(self) -> tuple[str, str, str]:
        return (self.preset_label, self.fixture_id, self.preset_no)

    @property
    def file_stem(self) -> str:
        return safe_filename(f"{self.preset_no}_{self.fixture_id}")


@dataclass(frozen=True)
class PartituraRow:
    number: str
    name: str
    fade: str
    downfade: str
    delay: str
    trigger: str
    trigger_time: str
    info: str
    command: str

    def value(self, field_id: str) -> str:
        return {
            "number": self.number,
            "name": self.name,
            "fade": self.fade,
            "downfade": self.downfade,
            "delay": self.delay,
            "trigger": self.trigger,
            "trigger_time": self.trigger_time,
            "info": self.info,
            "command": self.command,
        }.get(field_id, "")


@dataclass
class PartituraField:
    field_id: str
    title: str
    enabled: bool


@dataclass(frozen=True)
class Project:
    title: str
    directory: Path
    xml_path: Path


@dataclass
class SftpConfig:
    host: str = ""
    port: int = 22
    username: str = ""
    password: str = ""
    remote_dir: str = REMOTE_PROJECT_ROOT_NAME


@dataclass(frozen=True)
class RemoteFile:
    remote_path: str
    relative_path: Path
    size: int = 0


PARTITURA_DEFAULT_FIELDS = [
    PartituraField("number", "Номер", True),
    PartituraField("name", "Реплика", True),
    PartituraField("trigger", "Trigger", False),
    PartituraField("trigger_time", "Trigger time", False),
    PartituraField("fade", "Fade", True),
    PartituraField("downfade", "Downfade", False),
    PartituraField("delay", "Delay", False),
    PartituraField("info", "Инфо", True),
    PartituraField("command", "Command", False),
]


def local_name(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]


def child_by_name(parent: ET.Element, name: str) -> Optional[ET.Element]:
    for child in parent:
        if local_name(child.tag) == name:
            return child
    return None


def children_by_name(parent: ET.Element, name: str) -> list[ET.Element]:
    return [child for child in parent if local_name(child.tag) == name]


def first_non_empty(*values: Optional[str]) -> str:
    for value in values:
        text = (value or "").strip()
        if text:
            return text
    return ""


def natural_key(value: str) -> tuple[int, int | str]:
    return (0, int(value)) if value.isdigit() else (1, value)


def safe_filename(value: str) -> str:
    value = re.sub(r"[^0-9A-Za-zА-Яа-яЁё_. -]+", "_", value, flags=re.UNICODE)
    value = re.sub(r"\s+", " ", value).strip(" ._")
    return value or "show"


def set_active_project_root(path: Path) -> None:
    global ACTIVE_PROJECT_ROOT
    ACTIVE_PROJECT_ROOT = path


def active_project_root() -> Path:
    return ACTIVE_PROJECT_ROOT


def project_dir_for_title(title: str) -> Path:
    return active_project_root() / safe_filename(title)


def display_title(value: str) -> str:
    text = value[:-9] if value.endswith("_passport") else value
    return text.replace("_", " ")


def project_title_from_dir(path: Path) -> str:
    return display_title(path.name)


def ensure_project_root() -> None:
    active_project_root().mkdir(parents=True, exist_ok=True)


def project_xml_path(project_dir: Path) -> Optional[Path]:
    xmls = sorted(path for path in project_dir.iterdir() if path.suffix.lower() in XML_EXTENSIONS and path.is_file())
    return xmls[0] if xmls else None


def require_project_xml(project_dir: Path) -> Path:
    xml_path = project_xml_path(project_dir)
    if xml_path is None:
        raise RuntimeError(f"В папке проекта нет XML: {project_dir}")
    return xml_path


def list_projects() -> list[Project]:
    ensure_project_root()
    projects: list[Project] = []
    for directory in sorted([path for path in active_project_root().iterdir() if path.is_dir()], key=lambda p: p.name.lower()):
        xml_path = project_xml_path(directory)
        if xml_path:
            projects.append(Project(display_title(directory.name), directory, xml_path))
    return projects


def copy_xml_to_project(source: Path, title: str) -> Project:
    ensure_project_root()
    project_dir = project_dir_for_title(title)
    project_dir.mkdir(parents=True, exist_ok=True)
    target = project_dir / f"{safe_filename(title)}.xml"
    if source.resolve() != target.resolve():
        shutil.copy2(source, target)
    return Project(display_title(title), project_dir, target)


def preset_xlsx_path(project_dir: Path, title: str) -> Path:
    return project_dir / f"{safe_filename(title)}_пресеты.xlsx"


def preset_pdf_path(project_dir: Path, title: str) -> Path:
    return project_dir / f"{safe_filename(title)}_пресеты.pdf"


def partitura_xlsx_path(project_dir: Path, title: str) -> Path:
    return project_dir / f"{safe_filename(title)}_партитура.xlsx"


def partitura_pdf_path(project_dir: Path, title: str) -> Path:
    return project_dir / f"{safe_filename(title)}_партитура.pdf"


def photos_dir(project_dir: Path) -> Path:
    path = project_dir / "photos"
    path.mkdir(parents=True, exist_ok=True)
    return path


def find_existing_presets_xlsx(project_dir: Path) -> Optional[Path]:
    candidates = sorted(path for path in project_dir.glob("*_пресеты.xlsx") if not path.name.startswith("._"))
    if candidates:
        return candidates[0]
    legacy = project_dir / "passport.xlsx"
    return legacy if legacy.exists() else None


def load_sftp_config() -> SftpConfig:
    try:
        data = json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return SftpConfig()
    raw_host = str(data.get("host", data.get("url", ""))).strip()
    raw_host = raw_host.removeprefix("sftp://").removeprefix("ssh://")
    raw_host = re.sub(r"^https?://", "", raw_host).split("/", 1)[0]
    try:
        port = int(data.get("port", 22))
    except (TypeError, ValueError):
        port = 22
    return SftpConfig(
        host=raw_host,
        port=port,
        username=str(data.get("username", "")),
        password=str(data.get("password", "")),
        remote_dir=str(data.get("remote_dir", REMOTE_PROJECT_ROOT_NAME)) or REMOTE_PROJECT_ROOT_NAME,
    )


def save_sftp_config(config: SftpConfig) -> None:
    APP_DATA_DIR.mkdir(parents=True, exist_ok=True)
    CONFIG_PATH.write_text(
        json.dumps(
            {
                "host": config.host,
                "port": config.port,
                "username": config.username,
                "password": config.password,
                "remote_dir": config.remote_dir,
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )


def local_project_files(local_dir: Path) -> list[Path]:
    return sorted(
        [path for path in local_dir.rglob("*") if path.is_file() and not path.name.endswith(".download")],
        key=lambda path: (0 if path.suffix.lower() in XML_EXTENSIONS else 1, str(path.relative_to(local_dir)).lower()),
    )


def sftp_join(*parts: str) -> str:
    cleaned = []
    absolute = False
    for part in parts:
        if not part:
            continue
        text = str(part).replace("\\", "/")
        if text.startswith("/") and not cleaned:
            absolute = True
        cleaned.extend(piece for piece in text.split("/") if piece)
    prefix = "/" if absolute else ""
    return prefix + "/".join(cleaned)


def sftp_parent(path: str) -> str:
    clean = path.rstrip("/")
    if "/" not in clean:
        return ""
    return clean.rsplit("/", 1)[0]


def sftp_basename(path: str) -> str:
    return path.rstrip("/").rsplit("/", 1)[-1]


def sftp_exists(sftp: paramiko.SFTPClient, path: str) -> bool:
    try:
        sftp.stat(path)
        return True
    except FileNotFoundError:
        return False
    except OSError:
        return False


def ensure_sftp_dir(sftp: paramiko.SFTPClient, path: str) -> None:
    current = "/" if path.startswith("/") else ""
    for part in [piece for piece in path.split("/") if piece]:
        current = sftp_join(current, part)
        try:
            sftp.mkdir(current)
        except OSError:
            pass


def sftp_is_dir(sftp: paramiko.SFTPClient, path: str) -> bool:
    try:
        return stat.S_ISDIR(sftp.stat(path).st_mode)
    except OSError:
        return False


def remove_sftp_path(sftp: paramiko.SFTPClient, path: str) -> None:
    if not sftp_exists(sftp, path):
        return
    if sftp_is_dir(sftp, path):
        for item in sftp.listdir_attr(path):
            remove_sftp_path(sftp, sftp_join(path, item.filename))
        sftp.rmdir(path)
    else:
        sftp.remove(path)


def sftp_project_names(sftp: paramiko.SFTPClient, remote_root: str) -> list[str]:
    ensure_sftp_dir(sftp, remote_root)
    names: list[str] = []
    for item in sftp.listdir_attr(remote_root):
        path = sftp_join(remote_root, item.filename)
        if stat.S_ISDIR(item.st_mode):
            try:
                if any(name.lower().endswith(".xml") for name in sftp.listdir(path)):
                    names.append(item.filename)
            except OSError:
                pass
    return sorted(names, key=str.lower)


def download_sftp_project_index(sftp: paramiko.SFTPClient, remote_root: str, local_root: Path) -> None:
    if local_root.exists():
        shutil.rmtree(local_root)
    local_root.mkdir(parents=True, exist_ok=True)
    for project_name in sftp_project_names(sftp, remote_root):
        remote_project = sftp_join(remote_root, project_name)
        local_project = local_root / project_name
        local_project.mkdir(parents=True, exist_ok=True)
        for item in sftp.listdir_attr(remote_project):
            if stat.S_ISDIR(item.st_mode):
                continue
            sftp.get(sftp_join(remote_project, item.filename), str(local_project / item.filename))


def collect_sftp_files(sftp: paramiko.SFTPClient, remote_dir: str, base_relative: Path = Path()) -> list[RemoteFile]:
    files: list[RemoteFile] = []
    for item in sftp.listdir_attr(remote_dir):
        remote_path = sftp_join(remote_dir, item.filename)
        relative = base_relative / item.filename
        if stat.S_ISDIR(item.st_mode):
            files.extend(collect_sftp_files(sftp, remote_path, relative))
        else:
            files.append(RemoteFile(remote_path, relative, int(item.st_size or 0)))
    return files


def download_sftp_project_atomic(
    sftp: paramiko.SFTPClient,
    remote_project: str,
    local_project: Path,
    progress: Optional[Callable[[int, int, str], None]] = None,
) -> None:
    if progress:
        progress(0, 0, "собираю список файлов")
    files = collect_sftp_files(sftp, remote_project)
    temp_dir = local_project.parent / f".{local_project.name}.download"
    backup_dir = local_project.parent / f".{local_project.name}.old"
    if temp_dir.exists():
        shutil.rmtree(temp_dir)
    temp_dir.mkdir(parents=True, exist_ok=True)
    try:
        total = len(files)
        for index, remote_file in enumerate(files, start=1):
            if progress:
                progress(index, total, str(remote_file.relative_path))
            local_path = temp_dir / remote_file.relative_path
            local_path.parent.mkdir(parents=True, exist_ok=True)
            sftp.get(remote_file.remote_path, str(local_path))
        if not project_xml_path(temp_dir):
            raise RuntimeError("В скачанном проекте нет XML.")
        if backup_dir.exists():
            shutil.rmtree(backup_dir)
        if local_project.exists():
            local_project.replace(backup_dir)
        temp_dir.replace(local_project)
        if backup_dir.exists():
            shutil.rmtree(backup_dir)
    except Exception:
        if temp_dir.exists():
            shutil.rmtree(temp_dir)
        raise


def upload_sftp_project_atomic(
    sftp: paramiko.SFTPClient,
    local_project: Path,
    remote_project: str,
    progress: Optional[Callable[[int, int, str], None]] = None,
) -> None:
    require_project_xml(local_project)
    remote_root = sftp_parent(remote_project)
    ensure_sftp_dir(sftp, remote_root)
    temp_remote = remote_project.rstrip("/") + ".upload"
    remove_sftp_path(sftp, temp_remote)
    ensure_sftp_dir(sftp, temp_remote)
    try:
        files = local_project_files(local_project)
        total = len(files)
        for index, local_path in enumerate(files, start=1):
            relative = local_path.relative_to(local_project)
            if progress:
                progress(index, total, str(relative))
            remote_parent = temp_remote
            for part in relative.parts[:-1]:
                remote_parent = sftp_join(remote_parent, part)
                ensure_sftp_dir(sftp, remote_parent)
            sftp.put(str(local_path), sftp_join(remote_parent, relative.name))
        remove_sftp_path(sftp, remote_project)
        try:
            sftp.rename(temp_remote, remote_project)
        except OSError:
            ensure_sftp_dir(sftp, sftp_parent(remote_project))
            sftp.rename(temp_remote, remote_project)
    except Exception:
        remove_sftp_path(sftp, temp_remote)
        raise


def copy_project_dir(source: Path, target_root: Path, replace: bool) -> Path:
    target = target_root / source.name
    if target.exists():
        if not replace:
            raise FileExistsError(target)
        shutil.rmtree(target)
    target_root.mkdir(parents=True, exist_ok=True)
    shutil.copytree(source, target)
    return target


def mirror_project_to_remote_cache(project_dir: Path) -> None:
    target = REMOTE_CACHE_ROOT / project_dir.name
    try:
        if target.resolve() == project_dir.resolve():
            return
    except OSError:
        pass
    if target.exists():
        shutil.rmtree(target)
    REMOTE_CACHE_ROOT.mkdir(parents=True, exist_ok=True)
    shutil.copytree(project_dir, target)


def parse_grandma2_presets(xml_path: Path) -> list[PresetItem]:
    tree = ET.parse(xml_path)
    root = tree.getroot()
    ordered: dict[tuple[str, str], set[str]] = {}

    for cue_data in root.iter():
        if local_name(cue_data.tag) != "CueData":
            continue

        channel = child_by_name(cue_data, "Channel")
        preset = child_by_name(cue_data, "Preset")
        if channel is None or preset is None:
            continue

        fixture_id = channel.get("fixture_id") or channel.get("channel_id")
        preset_name = (preset.get("name") or "").strip()
        no_parts = [
            (node.text or "").strip()
            for node in children_by_name(preset, "No")
            if (node.text or "").strip()
        ]
        preset_no = ".".join(no_parts)
        if not fixture_id or not preset_name or not preset_no:
            continue

        ordered.setdefault((preset_no, preset_name), set()).add(fixture_id)

    result: list[PresetItem] = []
    for (preset_no, preset_name), fixture_ids in ordered.items():
        for fixture_id in sorted(fixture_ids, key=natural_key):
            result.append(PresetItem(preset_no, preset_name, fixture_id))
    return result


def format_cue_number(number: str, sub_number: str) -> str:
    if not sub_number or sub_number == "0":
        return number
    try:
        value = int(sub_number)
        if value % 100 == 0:
            return f"{number}.{value // 100}"
        if value % 10 == 0:
            return f"{number}.{value // 10}"
        return f"{number}.{value}"
    except ValueError:
        return f"{number}.{sub_number}"


def split_cue_number(value: str) -> tuple[str, str]:
    text = str(value or "").strip()
    if not text:
        return "0", "0"
    if "." not in text:
        return text, "0"
    number, sub = text.split(".", 1)
    sub = re.sub(r"\D", "", sub)
    if not sub:
        return number or "0", "0"
    return number or "0", str(int(sub) * 100)


def cue_xml_key(cue_number: str, order: int) -> str:
    return f"{cue_number}#{order}"


def iter_partitura_cues(root: ET.Element) -> list[tuple[ET.Element, str]]:
    cues: list[tuple[ET.Element, str]] = []
    order = 0
    for cue in root.iter():
        if local_name(cue.tag) != "Cue":
            continue
        number_node = child_by_name(cue, "Number")
        if number_node is None:
            continue
        cue_part = first_index_zero_cue_part(cue)
        if cue_part is None:
            continue
        cue_number = format_cue_number(number_node.get("number", ""), number_node.get("sub_number", "0"))
        cues.append((cue, cue_xml_key(cue_number, order)))
        order += 1
    return cues


def first_index_zero_cue_part(cue: ET.Element) -> Optional[ET.Element]:
    for cue_part in cue.iter():
        if local_name(cue_part.tag) == "CuePart" and (cue_part.get("index") or "0") == "0":
            return cue_part
    return None


def partitura_xml_keys(xml_path: Path) -> list[str]:
    tree = ET.parse(xml_path)
    return [key for _cue, key in iter_partitura_cues(tree.getroot())]


def parse_partitura(xml_path: Path) -> list[PartituraRow]:
    tree = ET.parse(xml_path)
    root = tree.getroot()
    rows: list[PartituraRow] = []

    for cue in root.iter():
        if local_name(cue.tag) != "Cue":
            continue

        number_node = child_by_name(cue, "Number")
        if number_node is None:
            continue

        cue_number = format_cue_number(number_node.get("number", ""), number_node.get("sub_number", "0"))
        trigger_node = child_by_name(cue, "Trigger")
        trigger = trigger_node.get("type") if trigger_node is not None else "Go"
        trigger_time = trigger_node.get("data_f", "") if trigger_node is not None else ""
        cue_info = ""
        cue_command = ""

        info_items = child_by_name(cue, "InfoItems")
        if info_items is not None:
            info_node = child_by_name(info_items, "Info")
            if info_node is not None and info_node.text:
                cue_info = info_node.text.strip()

        for child in cue:
            if local_name(child.tag) in {"Command", "Cmd", "CueCommand", "CLI", "CommandLine"}:
                cue_command = first_non_empty(
                    child.get("command"),
                    child.get("cmd"),
                    child.get("command_text"),
                    child.text,
                    cue_command,
                )

        for cue_part in cue.iter():
            if local_name(cue_part.tag) != "CuePart":
                continue
            if (cue_part.get("index") or "0") != "0":
                continue

            part_info = ""
            part_command = ""
            for child in cue_part.iter():
                if child is cue_part:
                    continue
                child_name = local_name(child.tag)
                if child_name == "Info" and child.text:
                    part_info = child.text.strip()
                elif child_name in {"Command", "Cmd", "CueCommand", "CLI", "CommandLine"}:
                    part_command = first_non_empty(
                        child.get("command"),
                        child.get("cmd"),
                        child.get("command_text"),
                        child.text,
                        part_command,
                    )

            rows.append(
                PartituraRow(
                    number=cue_number,
                    name=cue_part.get("name", "cue"),
                    fade=cue_part.get("basic_fade", "0"),
                    downfade=cue_part.get("basic_downfade", ""),
                    delay=cue_part.get("basic_delay", ""),
                    trigger=trigger or "Go",
                    trigger_time=trigger_time,
                    info=part_info or cue_info,
                    command=part_command or cue_command,
                )
            )
    return rows


def save_partitura_to_show_xml(xml_path: Path, rows: list[dict], output_path: Path) -> None:
    parser = ET.XMLParser(encoding="utf-8")
    tree = ET.parse(xml_path, parser=parser)
    root = tree.getroot()
    namespace = root.tag.split("}", 1)[0][1:] if root.tag.startswith("{") else ""
    if namespace:
        ET.register_namespace("", namespace)
    xsi_namespace = root.attrib.get("{http://www.w3.org/2001/XMLSchema-instance}schemaLocation")
    if xsi_namespace is not None:
        ET.register_namespace("xsi", "http://www.w3.org/2001/XMLSchema-instance")

    parent_map = {child: parent for parent in root.iter() for child in parent}
    cues_by_key = {key: cue for cue, key in iter_partitura_cues(root)}
    all_cues_by_key = dict(cues_by_key)
    used_keys: set[str] = set()
    template_cue = next((cue for cue, _key in iter_partitura_cues(root)), None)

    for row_index, row in enumerate(rows):
        key = row.get("_xml_key") or f"__new_{row_index}"
        cue = cues_by_key.get(row.get("_xml_key") or "")
        if cue is None:
            cue = create_empty_cue(root, template_cue)
            insert_cue_for_row(root, parent_map, cue, row_index, rows, all_cues_by_key)
            parent_map[cue] = find_cue_parent(root)
            all_cues_by_key[key] = cue
        else:
            used_keys.add(row.get("_xml_key") or "")
        update_cue_from_partitura_row(cue, row)

    for key, cue in cues_by_key.items():
        if key in used_keys:
            continue
        parent = parent_map.get(cue)
        if parent is not None:
            parent.remove(cue)

    renumber_cue_indices(root)
    ET.indent(tree, space="\t")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    write_xml_preserving_preamble(tree, xml_path, output_path)


def write_xml_preserving_preamble(tree: ET.ElementTree, source_path: Path, output_path: Path) -> None:
    buffer = io.BytesIO()
    tree.write(buffer, encoding="utf-8", xml_declaration=True)
    content = buffer.getvalue()
    preamble = []
    try:
        for raw_line in source_path.read_text(encoding="utf-8-sig").splitlines()[:10]:
            stripped = raw_line.strip()
            if stripped.startswith("<?xml-stylesheet"):
                preamble.append(stripped)
    except OSError:
        preamble = []
    if not preamble:
        output_path.write_bytes(content)
        return
    lines = content.splitlines(keepends=True)
    insert = "".join(f"{line}\n" for line in preamble).encode("utf-8")
    if lines:
        output_path.write_bytes(lines[0] + insert + b"".join(lines[1:]))
    else:
        output_path.write_bytes(insert + content)


def create_empty_cue(root: ET.Element, template_cue: Optional[ET.Element]) -> ET.Element:
    tag = qname(root, "Cue")
    cue = ET.Element(tag)
    cue.append(ET.Element(qname(root, "Number"), {"number": "0", "sub_number": "0"}))
    cue.append(ET.Element(qname(root, "CueDatas")))
    cue_part = ET.Element(qname(root, "CuePart"), {"index": "0", "name": "cue", "basic_fade": "0"})
    timing = ET.SubElement(cue_part, qname(root, "CuePartPresetTiming"))
    for _ in range(10):
        ET.SubElement(timing, qname(root, "PresetTiming"))
    cue.append(cue_part)
    if template_cue is not None and "index" in template_cue.attrib:
        cue.set("index", template_cue.get("index", "1"))
    return cue


def qname(root: ET.Element, name: str) -> str:
    if root.tag.startswith("{"):
        namespace = root.tag.split("}", 1)[0][1:]
        return f"{{{namespace}}}{name}"
    return name


def find_cue_parent(root: ET.Element) -> ET.Element:
    for parent in root.iter():
        if any(local_name(child.tag) == "Cue" for child in parent):
            return parent
    return root


def insert_cue_for_row(root: ET.Element, parent_map: dict[ET.Element, ET.Element], cue: ET.Element, row_index: int, rows: list[dict], cues_by_key: dict[str, ET.Element]) -> None:
    parent = find_cue_parent(root)
    insert_after: Optional[ET.Element] = None
    for previous_index in range(row_index - 1, -1, -1):
        previous = rows[previous_index]
        previous_key = previous.get("_xml_key") or f"__new_{previous_index}"
        if previous_key in cues_by_key:
            insert_after = cues_by_key[previous_key]
            break
    if insert_after is not None and parent_map.get(insert_after) is parent:
        parent.insert(list(parent).index(insert_after) + 1, cue)
    else:
        cue_children = [child for child in parent if local_name(child.tag) == "Cue"]
        if cue_children:
            parent.insert(list(parent).index(cue_children[-1]) + 1, cue)
        else:
            parent.append(cue)


def update_cue_from_partitura_row(cue: ET.Element, row: dict) -> None:
    number_node = child_by_name(cue, "Number")
    if number_node is None:
        number_node = ET.Element(qname(cue, "Number"))
        cue.insert(0, number_node)
    number, sub_number = split_cue_number(row.get("number", ""))
    number_node.set("number", number)
    number_node.set("sub_number", sub_number)

    cue_part = first_index_zero_cue_part(cue)
    if cue_part is None:
        cue_part = ET.SubElement(cue, qname(cue, "CuePart"), {"index": "0"})
    cue_part.set("index", "0")
    cue_part.set("name", str(row.get("name", "") or "cue"))
    set_optional_attr(cue_part, "basic_fade", row.get("fade"), default="0")
    set_optional_attr(cue_part, "basic_downfade", row.get("downfade"))
    set_optional_attr(cue_part, "basic_delay", row.get("delay"))

    set_trigger(cue, row)
    set_info(cue, row.get("info", ""))
    set_command(cue, row.get("command", ""))


def set_optional_attr(element: ET.Element, name: str, value: object, default: str | None = None) -> None:
    text = str(value or "").strip()
    if text:
        element.set(name, text)
    elif default is not None:
        element.set(name, default)
    else:
        element.attrib.pop(name, None)


def set_trigger(cue: ET.Element, row: dict) -> None:
    trigger_type = str(row.get("trigger", "") or "Go").strip() or "Go"
    trigger_time = str(row.get("trigger_time", "") or "").strip()
    trigger_node = child_by_name(cue, "Trigger")
    if trigger_type == "Go" and not trigger_time:
        if trigger_node is not None:
            cue.remove(trigger_node)
        return
    if trigger_node is None:
        number_node = child_by_name(cue, "Number")
        trigger_node = ET.Element(qname(cue, "Trigger"))
        cue.insert((list(cue).index(number_node) + 1) if number_node is not None else 0, trigger_node)
    trigger_node.set("type", trigger_type)
    if trigger_time:
        trigger_node.set("data_f", trigger_time)
    else:
        trigger_node.attrib.pop("data_f", None)


def set_info(cue: ET.Element, value: object) -> None:
    text = str(value or "").strip()
    info_items = child_by_name(cue, "InfoItems")
    if not text:
        if info_items is not None:
            cue.remove(info_items)
        return
    if info_items is None:
        info_items = ET.Element(qname(cue, "InfoItems"))
        cue.insert(1, info_items)
    info_node = child_by_name(info_items, "Info")
    if info_node is None:
        info_node = ET.SubElement(info_items, qname(cue, "Info"))
    info_node.text = text


def set_command(cue: ET.Element, value: object) -> None:
    text = str(value or "").strip()
    command_node = None
    for child in cue:
        if local_name(child.tag) in {"Command", "Cmd", "CueCommand", "CLI", "CommandLine"}:
            command_node = child
            break
    if not text:
        if command_node is not None:
            cue.remove(command_node)
        return
    if command_node is None:
        command_node = ET.Element(qname(cue, "Command"))
        cue.append(command_node)
    command_node.text = text


def renumber_cue_indices(root: ET.Element) -> None:
    for parent in root.iter():
        cue_index = 1
        for child in parent:
            if local_name(child.tag) != "Cue":
                continue
            if child.get("{http://www.w3.org/2001/XMLSchema-instance}nil") == "true":
                continue
            if child_by_name(child, "Number") is None:
                continue
            child.set("index", str(cue_index))
            cue_index += 1


def find_photo_for_stem(stem: str, photo_dir: Path) -> Optional[Path]:
    for extension in [".jpg", ".jpeg", ".png", ".webp", ".bmp"]:
        exact = photo_dir / f"{stem}{extension}"
        if exact.exists():
            return exact
    matches = sorted(
        path
        for path in photo_dir.glob(f"{stem}_*")
        if path.suffix.lower() in PHOTO_EXTENSIONS and path.is_file()
    )
    return matches[0] if matches else None


def read_passport_rows_from_sheet(sheet, item_by_group: dict[tuple[str, str], PresetItem]) -> list[PassportRow]:
    header_row = 2 if sheet.max_row >= 2 and sheet.cell(2, 1).value == "Пресет" else 1
    if sheet.cell(header_row, 1).value != "Пресет":
        return []

    rows: list[PassportRow] = []
    last_preset = ""
    last_fixture = ""
    for values in sheet.iter_rows(min_row=header_row + 1, values_only=True):
        preset, fixture, _photo, description = (list(values) + [None, None, None, None])[:4]
        preset_text = str(preset).strip() if preset is not None else last_preset
        fixture_text = str(fixture).strip() if fixture is not None else last_fixture
        description_text = "" if description is None else str(description)
        if not preset_text and not fixture_text and not description_text.strip():
            continue
        last_preset = preset_text
        last_fixture = fixture_text
        item = item_by_group.get((preset_text, fixture_text))
        preset_no = item.preset_no if item else ""
        rows.append(PassportRow(preset_text, fixture_text, preset_no, None, description_text))
    return rows


def load_passport_rows(items: list[PresetItem], project_dir: Path, title: str) -> tuple[list[PassportRow], Optional[str]]:
    item_by_group = {(item.preset_label, item.fixture_id): item for item in items}
    photo_dir = photos_dir(project_dir)
    xlsx_path = find_existing_presets_xlsx(project_dir)
    warning = None
    rows: list[PassportRow] = []

    if xlsx_path is not None:
        try:
            workbook = load_workbook(xlsx_path, data_only=True)
            rows = read_passport_rows_from_sheet(workbook.active, item_by_group)
        except (BadZipFile, OSError, ValueError) as exc:
            warning = f"Не удалось прочитать таблицу {xlsx_path.name}: {exc}"

    if not rows:
        rows = [PassportRow(item.preset_label, item.fixture_id, item.preset_no) for item in items]

    existing_groups = {(row.preset_label, row.fixture_id) for row in rows}
    for item in items:
        if (item.preset_label, item.fixture_id) not in existing_groups:
            rows.append(PassportRow(item.preset_label, item.fixture_id, item.preset_no))

    counters: dict[str, int] = {}
    for row in rows:
        if not row.preset_no:
            item = item_by_group.get((row.preset_label, row.fixture_id))
            row.preset_no = item.preset_no if item else ""
            
        # If the row already has a valid photo_path from JSON, keep it!
        if row.photo_path and row.photo_path.exists():
            continue

        stem = row.file_stem
        counters[stem] = counters.get(stem, 0) + 1
        stems = [stem]
        legacy_stem = safe_filename(f"{row.preset_label}_{row.fixture_id}")
        if legacy_stem not in stems:
            stems.append(legacy_stem)
            
        if counters[stem] == 1:
            row.photo_path = next((photo for candidate in stems if (photo := find_photo_for_stem(candidate, photo_dir))), None)
        else:
            row.photo_path = next((photo for candidate in stems if (photo := find_photo_for_stem(f"{candidate}_{counters[stem]}", photo_dir))), None)
    return rows, warning


def source_from_text(value: str):
    value = value.strip().split()[0] if value.strip() else "0"
    return int(value) if value.isdigit() else value





def create_cell_photo(source: Path, target: Path, width: int, height: int) -> None:
    with Image.open(source) as image:
        image = image.convert("RGB")
        image.thumbnail((width, height))
        canvas = Image.new("RGB", (width, height), "white")
        left = (width - image.width) // 2
        top = (height - image.height) // 2
        canvas.paste(image, (left, top))
        canvas.save(target)


def create_passport_xlsx(rows: list[PassportRow], path: Path, title: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Паспорт"
    ws.append([title])
    ws.merge_cells("A1:D1")
    ws.append(["Пресет", "Прибор", "Фото", "Описание"])

    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws["A1"].font = Font(bold=True, size=16)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30
    for cell in ws[2]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        cell.fill = PatternFill("solid", fgColor="EEEEEE")

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 42
    ws.column_dimensions["D"].width = 34
    photo_width = 300
    photo_height = 180

    with tempfile.TemporaryDirectory() as temp_dir:
        prepared_photos: list[Path] = []
        for row_index, row in enumerate(rows, start=3):
            ws.cell(row=row_index, column=1, value=row.preset_label)
            ws.cell(row=row_index, column=2, value=row.fixture_id)
            ws.cell(row=row_index, column=4, value=row.description)
            ws.row_dimensions[row_index].height = 138
            for col in range(1, 5):
                cell = ws.cell(row=row_index, column=col)
                cell.border = border
                cell.alignment = Alignment(horizontal="center" if col != 4 else "left", vertical="center", wrap_text=True)
            if row.photo_path and row.photo_path.exists():
                prepared = Path(temp_dir) / f"photo_{row_index}.png"
                create_cell_photo(row.photo_path, prepared, photo_width, photo_height)
                prepared_photos.append(prepared)
                img = XlsxImage(str(prepared))
                img.width = photo_width
                img.height = photo_height
                ws.add_image(img, f"C{row_index}")

        merge_duplicate_passport_cells(ws, rows)
        ws.freeze_panes = "A3"
        wb.save(path)


def merge_duplicate_passport_cells(ws, rows: list[PassportRow]) -> None:
    start = 3
    while start < len(rows) + 3:
        key = rows[start - 3].group_key
        end = start
        while end + 1 < len(rows) + 3 and rows[end - 2].group_key == key:
            end += 1
        if end > start:
            ws.merge_cells(start_row=start, start_column=1, end_row=end, end_column=1)
            ws.merge_cells(start_row=start, start_column=2, end_row=end, end_column=2)
        start = end + 1


def create_partitura_xlsx(rows: list[PartituraRow], fields: list[PartituraField], path: Path, title: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Партитура"
    last_col = max(1, len(fields))
    ws.append([title])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    ws.append([field.title for field in fields])

    thin = Side(style="thin", color="555555")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws["A1"].font = Font(bold=True, size=16)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    for cell in ws[2]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    widths = {
        "number": 12,
        "name": 48,
        "trigger": 13,
        "trigger_time": 15,
        "fade": 10,
        "downfade": 12,
        "delay": 10,
        "info": 54,
        "command": 32,
    }
    for index, field in enumerate(fields, start=1):
        ws.column_dimensions[column_letter(index)].width = widths.get(field.field_id, 18)

    for row in rows:
        ws.append([row.value(field.field_id) for field in fields])

    for sheet_row in ws.iter_rows(min_row=3):
        for col_index, cell in enumerate(sheet_row, start=1):
            field = fields[col_index - 1]
            align = "left" if field.field_id in {"name", "info", "command"} else "center"
            cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
            cell.border = border

    ws.freeze_panes = "A3"
    wb.save(path)




def create_passport_pdf(rows: list[PassportRow], path: Path, title: str) -> None:
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import landscape, A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.platypus import Image as PdfImage
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except ImportError:
        create_simple_pdf_placeholder(path, title, "Установите reportlab для PDF.")
        return

    font_name, bold_font_name = register_pdf_fonts()
    doc = SimpleDocTemplate(str(path), pagesize=landscape(A4), rightMargin=12 * mm, leftMargin=12 * mm, topMargin=10 * mm, bottomMargin=10 * mm)
    styles = getSampleStyleSheet()
    styles["Title"].fontName = bold_font_name
    normal = ParagraphStyle("normal", parent=styles["Normal"], fontName=font_name, fontSize=8, leading=10, alignment=1)
    left = ParagraphStyle("left", parent=normal, alignment=0)
    story = [Paragraph(title, styles["Title"]), Spacer(1, 5 * mm)]
    data = [["Пресет", "Прибор", "Фото", "Описание"]]
    temp_files: list[Path] = []
    for row in rows:
        photo = ""
        if row.photo_path and row.photo_path.exists():
            prepared = Path(tempfile.gettempdir()) / f"passport_pdf_{time.time_ns()}.jpg"
            create_cell_photo(row.photo_path, prepared, 240, 145)
            temp_files.append(prepared)
            photo = PdfImage(str(prepared), width=64 * mm, height=38 * mm)
        data.append([Paragraph(row.preset_label, normal), Paragraph(row.fixture_id, normal), photo, Paragraph(row.description, left)])
    table = Table(data, colWidths=[62 * mm, 22 * mm, 70 * mm, 90 * mm], repeatRows=1)
    style_commands = [
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("BACKGROUND", (0, 0), (-1, 0), colors.whitesmoke),
        ("FONTNAME", (0, 0), (-1, 0), bold_font_name),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]
    start = 1
    while start <= len(rows):
        key = rows[start - 1].group_key
        end = start
        while end < len(rows) and rows[end].group_key == key:
            end += 1
        if end > start:
            style_commands.append(("SPAN", (0, start), (0, end)))
            style_commands.append(("SPAN", (1, start), (1, end)))
        start = end + 1
    table.setStyle(TableStyle(style_commands))
    story.append(table)
    doc.build(story)
    for temp in temp_files:
        temp.unlink(missing_ok=True)


def create_partitura_pdf(rows: list[PartituraRow], fields: list[PartituraField], path: Path, title: str) -> None:
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except ImportError:
        create_simple_pdf_placeholder(path, title, "Установите reportlab для PDF.")
        return

    font_name, bold_font_name = register_pdf_fonts()
    doc = SimpleDocTemplate(str(path), pagesize=A4, rightMargin=10 * mm, leftMargin=10 * mm, topMargin=10 * mm, bottomMargin=10 * mm)
    styles = getSampleStyleSheet()
    styles["Title"].fontName = bold_font_name
    small = ParagraphStyle("small", parent=styles["Normal"], fontName=font_name, fontSize=6.4, leading=7.4)
    header = ParagraphStyle("header", parent=small, fontName=bold_font_name, alignment=1)
    story = [Paragraph(title, styles["Title"]), Spacer(1, 4 * mm)]
    data = [[Paragraph(field.title, header) for field in fields]]
    for row in rows:
        data.append([Paragraph(row.value(field.field_id).replace("\n", "<br/>"), small) for field in fields])
    weights = [partitura_pdf_weight(field.field_id) for field in fields]
    total = sum(weights) or 1
    width = 190 * mm
    col_widths = [width * weight / total for weight in weights]
    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.45, colors.black),
        ("BACKGROUND", (0, 0), (-1, 0), colors.white),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    story.append(table)
    doc.build(story)


def partitura_pdf_weight(field_id: str) -> float:
    return {
        "number": 0.75,
        "name": 2.4,
        "trigger": 0.9,
        "trigger_time": 1.05,
        "fade": 0.65,
        "downfade": 0.8,
        "delay": 0.65,
        "info": 2.6,
        "command": 1.4,
    }.get(field_id, 1.0)


def register_pdf_fonts() -> tuple[str, str]:
    try:
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
    except ImportError:
        return "Helvetica", "Helvetica-Bold"

    candidates = [
        (
            Path("/System/Library/Fonts/Supplemental/Arial.ttf"),
            Path("/System/Library/Fonts/Supplemental/Arial Bold.ttf"),
        ),
        (
            Path("/Library/Fonts/Arial.ttf"),
            Path("/Library/Fonts/Arial Bold.ttf"),
        ),
        (
            Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
            Path("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"),
        ),
    ]
    for regular, bold in candidates:
        if regular.exists() and bold.exists():
            try:
                if "PassportArial" not in pdfmetrics.getRegisteredFontNames():
                    pdfmetrics.registerFont(TTFont("PassportArial", str(regular)))
                if "PassportArialBold" not in pdfmetrics.getRegisteredFontNames():
                    pdfmetrics.registerFont(TTFont("PassportArialBold", str(bold)))
                return "PassportArial", "PassportArialBold"
            except Exception:
                continue
    return "Helvetica", "Helvetica-Bold"


def create_simple_pdf_placeholder(path: Path, title: str, message: str) -> None:
    image = Image.new("RGB", (1240, 1754), "white")
    draw = ImageDraw.Draw(image)
    font = ImageFont.load_default()
    draw.text((80, 80), title, fill="black", font=font)
    draw.text((80, 130), message, fill="black", font=font)
    image.save(path, "PDF")


def open_path(path: Path) -> None:
    if sys.platform == "darwin":
        subprocess.run(["open", str(path)], check=False)
    elif os.name == "nt":
        os.startfile(path)  # type: ignore[attr-defined]
    else:
        subprocess.run(["xdg-open", str(path)], check=False)


def reveal_path(path: Path) -> None:
    if sys.platform == "darwin":
        subprocess.run(["open", "-R", str(path)], check=False)
    elif os.name == "nt":
        subprocess.run(["explorer", f"/select,{path}"], check=False)
    else:
        open_path(path.parent)


def print_xml_summary(xml_path: Path) -> None:
    items = parse_grandma2_presets(xml_path)
    print(f"XML: {xml_path}")
    print(f"Пресетов: {len({(item.preset_no, item.preset_name) for item in items})}")
    print(f"Строк пресет-прибор: {len(items)}")
    for item in items[:10]:
        print(f"  {item.preset_label} | {item.preset_no} | fixture {item.fixture_id}")
    if len(items) > 10:
        print("  ...")



if __name__ == "__main__":
    multiprocessing.freeze_support()
    if len(sys.argv) == 3 and sys.argv[1] == "--check":
        print_xml_summary(Path(sys.argv[2]))
    
